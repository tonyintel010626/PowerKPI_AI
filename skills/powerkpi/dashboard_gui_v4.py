#!/usr/bin/env python3
r"""
Enhanced GUI-based interactive dashboard generator with multiple graph types.

This script provides a graphical user interface for generating different types of
comparative visualizations from Power KPI workload data.

Supported Graph Types:
- Bar Chart: Compare metrics across workloads
- Line Graph: Trend analysis across workloads
- Scatter Plot: Distribution and correlation analysis
- Heatmap/Matrix: Compare all metrics across all workloads
- Box Plot: Statistical distribution comparison
- Grouped Bar: Side-by-side comparison

Usage:
    python dashboard_gui_v2.py
    python dashboard_gui_v2.py --folder C:\_hopper_results

Author: PowerKPI_Validator
Date: 2026-04-09
"""

import argparse
import json
import logging
import sys
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from pathlib import Path
from typing import Dict, List, Any, Optional, Set
import threading
import pandas as pd
import numpy as np
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import plotly.express as px
from nptdms import TdmsFile
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
# GENI integration removed - agent will handle this via MCP
# from geni_trend_analyzer import GENITrendAnalyzer
try:
    from hsdes_power_debug import HSDESPowerDebugger
    HSDES_AVAILABLE = True
except ImportError:
    HSDES_AVAILABLE = False
    logger.warning("HSDES power debugger not available - install pysvtools for HSDES integration")


# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class PowerKPIDashboardGUI:
    """Enhanced GUI application for Power KPI dashboard generation."""
    
    def __init__(self, root, initial_folder=None):
        """Initialize the GUI."""
        self.root = root
        self.root.title("Power KPI Dashboard Generator v3.4 - Multi-Instrument Graphs")
        self.root.geometry("1200x900")
        
        # Data storage
        self.results_folders = []
        self.all_power_rails = set()
        self.all_socwatch_metrics = set()
        self.all_workloads = set()
        self.parsed_data = None
        self.folder_name_map = {}  # Map workload names to full folder paths
        
        # UI variables
        self.folder_path_var = tk.StringVar()
        self.rail_filter_var = tk.StringVar()
        self.socwatch_filter_var = tk.StringVar()
        self.workload_filter_var = tk.StringVar()
        self.output_path_var = tk.StringVar(value="power_kpi_dashboard.html")
        self.graph_type_var = tk.StringVar(value="bar")
        self.metric_type_var = tk.StringVar(value="mean")
        self.data_source_var = tk.StringVar(value="results_json")  # NEW v3
        self.include_trends_var = tk.BooleanVar(value=True)  # NEW v3.2 - AI Trend Analysis
        self.query_hsdes_var = tk.BooleanVar(value=False)  # NEW v3.3 - HSDES Sighting Query
        self.project_release_var = tk.StringVar(value="")  # Project release filter for HSDES
        
        # NEW v3.4: Per-instrument graph type selection
        # DAQ/Power -> Scatter (best for showing distribution across rails)
        # SocWatch   -> Bar    (best for comparing C-state/freq metrics across workloads)
        # PerfTracer -> Bar    (grouped comparison of perf counters)
        # PowerTrace -> Bar    (comparison of trace aggregates)
        self.instrument_graph_types = {
            'daq': tk.StringVar(value="scatter"),
            'socwatch': tk.StringVar(value="bar"),
            'perftracer': tk.StringVar(value="bar"),
            'powertrace': tk.StringVar(value="bar")
        }
        self.instrument_enabled = {
            'daq': tk.BooleanVar(value=True),
            'socwatch': tk.BooleanVar(value=True),
            'perftracer': tk.BooleanVar(value=False),
            'powertrace': tk.BooleanVar(value=False)
        }
        
        # Checkbox variables
        self.rail_checkboxes = {}
        self.socwatch_checkboxes = {}
        self.workload_checkboxes = {}
        
        self.create_widgets()
        
        if initial_folder:
            self.folder_path_var.set(str(initial_folder))
            self.scan_folders()
    
    def create_widgets(self):
        """Create all GUI widgets."""
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(1, weight=1)
        
        # Title
        title = ttk.Label(main_frame, text="Power KPI Dashboard Generator v3.4", 
                         font=('Arial', 16, 'bold'))
        title.grid(row=0, column=0, pady=10)
        
        # Notebook for tabs
        notebook = ttk.Notebook(main_frame)
        notebook.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Tab 1: Folder Selection
        self.create_folder_tab(notebook)
        
        # Tab 2: Power Rails Selection
        self.create_rails_tab(notebook)
        
        # Tab 3: SocWatch Metrics Selection
        self.create_socwatch_tab(notebook)
        
        # Tab 4: Workload Selection
        self.create_workload_tab(notebook)
        
        # Tab 5: Graph Type & Generate
        self.create_output_tab(notebook)
        
        # Status bar
        self.status_var = tk.StringVar(value="Ready")
        status_bar = ttk.Label(main_frame, textvariable=self.status_var, 
                              relief=tk.SUNKEN, anchor=tk.W)
        status_bar.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=5)
    
    def create_folder_tab(self, notebook):
        """Create folder selection tab."""
        tab = ttk.Frame(notebook, padding="10")
        notebook.add(tab, text="1. Select Folders")
        
        ttk.Label(tab, text="Select the parent folder containing Hopper results:", 
                 font=('Arial', 10, 'bold')).grid(row=0, column=0, columnspan=3, pady=5, sticky=tk.W)
        
        ttk.Label(tab, text="Folder Path:").grid(row=1, column=0, sticky=tk.W, pady=5)
        ttk.Entry(tab, textvariable=self.folder_path_var, width=60).grid(row=1, column=1, pady=5)
        ttk.Button(tab, text="Browse...", command=self.browse_folder).grid(row=1, column=2, pady=5, padx=5)
        
        ttk.Button(tab, text="Scan Folders", command=self.scan_folders, 
                  style='Accent.TButton').grid(row=2, column=1, pady=10)
        
        ttk.Label(tab, text="Found Results Folders:", 
                 font=('Arial', 10, 'bold')).grid(row=3, column=0, columnspan=3, pady=5, sticky=tk.W)
        
        list_frame = ttk.Frame(tab)
        list_frame.grid(row=4, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        scrollbar = ttk.Scrollbar(list_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.folders_listbox = tk.Listbox(list_frame, yscrollcommand=scrollbar.set, 
                                         height=15, width=100)
        self.folders_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=self.folders_listbox.yview)
        
        tab.columnconfigure(1, weight=1)
        tab.rowconfigure(4, weight=1)
    
    def create_rails_tab(self, notebook):
        """Create power rails selection tab."""
        tab = ttk.Frame(notebook, padding="10")
        notebook.add(tab, text="2. Select Power Rails")
        
        filter_frame = ttk.LabelFrame(tab, text="Filter by Keyword", padding="5")
        filter_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=5)
        
        ttk.Label(filter_frame, text="Keyword:").grid(row=0, column=0, sticky=tk.W)
        filter_entry = ttk.Entry(filter_frame, textvariable=self.rail_filter_var, width=40)
        filter_entry.grid(row=0, column=1, padx=5)
        filter_entry.bind('<KeyRelease>', lambda e: self.filter_rails())
        
        ttk.Button(filter_frame, text="Clear", command=self.clear_rail_filter).grid(row=0, column=2, padx=5)
        ttk.Button(filter_frame, text="Select All Visible", 
                  command=self.select_all_rails).grid(row=0, column=3, padx=5)
        ttk.Button(filter_frame, text="Deselect All", 
                  command=self.deselect_all_rails).grid(row=0, column=4, padx=5)
        
        ttk.Label(tab, text="Available Power Rails (check to include in dashboard):", 
                 font=('Arial', 10, 'bold')).grid(row=1, column=0, pady=5, sticky=tk.W)
        
        checkbox_frame = ttk.Frame(tab)
        checkbox_frame.grid(row=2, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        canvas = tk.Canvas(checkbox_frame, height=400)
        scrollbar = ttk.Scrollbar(checkbox_frame, orient="vertical", command=canvas.yview)
        self.rails_checkbox_frame = ttk.Frame(canvas)
        
        self.rails_checkbox_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=self.rails_checkbox_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        tab.columnconfigure(0, weight=1)
        tab.rowconfigure(2, weight=1)
    
    def create_socwatch_tab(self, notebook):
        """Create SocWatch metrics selection tab."""
        tab = ttk.Frame(notebook, padding="10")
        notebook.add(tab, text="3. Select SocWatch Metrics")
        
        filter_frame = ttk.LabelFrame(tab, text="Filter by Keyword", padding="5")
        filter_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=5)
        
        ttk.Label(filter_frame, text="Keyword:").grid(row=0, column=0, sticky=tk.W)
        filter_entry = ttk.Entry(filter_frame, textvariable=self.socwatch_filter_var, width=40)
        filter_entry.grid(row=0, column=1, padx=5)
        filter_entry.bind('<KeyRelease>', lambda e: self.filter_socwatch())
        
        ttk.Button(filter_frame, text="Clear", command=self.clear_socwatch_filter).grid(row=0, column=2, padx=5)
        ttk.Button(filter_frame, text="Select All Visible", 
                  command=self.select_all_socwatch).grid(row=0, column=3, padx=5)
        ttk.Button(filter_frame, text="Deselect All", 
                  command=self.deselect_all_socwatch).grid(row=0, column=4, padx=5)
        
        ttk.Label(tab, text="Available SocWatch Metrics (check to include in dashboard):", 
                 font=('Arial', 10, 'bold')).grid(row=1, column=0, pady=5, sticky=tk.W)
        
        checkbox_frame = ttk.Frame(tab)
        checkbox_frame.grid(row=2, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        canvas = tk.Canvas(checkbox_frame, height=400)
        scrollbar = ttk.Scrollbar(checkbox_frame, orient="vertical", command=canvas.yview)
        self.socwatch_checkbox_frame = ttk.Frame(canvas)
        
        self.socwatch_checkbox_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=self.socwatch_checkbox_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        tab.columnconfigure(0, weight=1)
        tab.rowconfigure(2, weight=1)
    
    def create_workload_tab(self, notebook):
        """Create workload selection tab."""
        tab = ttk.Frame(notebook, padding="10")
        notebook.add(tab, text="4. Select Workloads")
        
        filter_frame = ttk.LabelFrame(tab, text="Filter by Keyword", padding="5")
        filter_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=5)
        
        ttk.Label(filter_frame, text="Keyword:").grid(row=0, column=0, sticky=tk.W)
        filter_entry = ttk.Entry(filter_frame, textvariable=self.workload_filter_var, width=40)
        filter_entry.grid(row=0, column=1, padx=5)
        filter_entry.bind('<KeyRelease>', lambda e: self.filter_workloads())
        
        ttk.Button(filter_frame, text="Clear", command=self.clear_workload_filter).grid(row=0, column=2, padx=5)
        ttk.Button(filter_frame, text="Select All Visible", 
                  command=self.select_all_workloads).grid(row=0, column=3, padx=5)
        ttk.Button(filter_frame, text="Deselect All", 
                  command=self.deselect_all_workloads).grid(row=0, column=4, padx=5)
        
        ttk.Label(tab, text="Available Workloads (check to include in dashboard):", 
                 font=('Arial', 10, 'bold')).grid(row=1, column=0, pady=5, sticky=tk.W)
        
        checkbox_frame = ttk.Frame(tab)
        checkbox_frame.grid(row=2, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        canvas = tk.Canvas(checkbox_frame, height=400)
        scrollbar = ttk.Scrollbar(checkbox_frame, orient="vertical", command=canvas.yview)
        self.workload_checkbox_frame = ttk.Frame(canvas)
        
        self.workload_checkbox_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=self.workload_checkbox_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        tab.columnconfigure(0, weight=1)
        tab.rowconfigure(2, weight=1)
    
    def create_output_tab(self, notebook):
        """Create output configuration tab with per-instrument graph selection."""
        tab = ttk.Frame(notebook, padding="10")
        notebook.add(tab, text="5. Graph Types & Generate")
        
        # Title
        ttk.Label(tab, text="Configure graphs for each instrument:", 
                 font=('Arial', 12, 'bold')).grid(row=0, column=0, columnspan=3, pady=10, sticky=tk.W)
        
        # V3.4: Per-Instrument Graph Configuration
        instruments_frame = ttk.LabelFrame(tab, text="Instrument Graph Configuration", padding="10")
        instruments_frame.grid(row=1, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=10)
        
        graph_type_options = [
            ("bar", "Bar Chart"),
            ("grouped_bar", "Grouped Bar"),
            ("line", "Line Graph"),
            ("scatter", "Scatter Plot"),
            ("heatmap", "Heatmap"),
            ("box", "Box Plot")
        ]
        
        instruments = [
            ('daq', 'DAQ (Power Rails)', 'Power rails from FlexLogger/DAQ'),
            ('socwatch', 'SocWatch', 'Package C-states, CPU states, frequencies'),
            ('perftracer', 'PerfTracer', 'Performance metrics and traces'),
            ('powertrace', 'PowerTrace', 'Power trace data and analysis')
        ]
        
        row_idx = 0
        for inst_id, inst_label, inst_desc in instruments:
            # Checkbox to enable/disable instrument
            enable_cb = ttk.Checkbutton(
                instruments_frame,
                text=f"☐ {inst_label}",
                variable=self.instrument_enabled[inst_id]
            )
            enable_cb.grid(row=row_idx, column=0, sticky=tk.W, padx=5, pady=5)
            
            # Description
            ttk.Label(instruments_frame, text=inst_desc, foreground="gray", font=('Arial', 8)).grid(
                row=row_idx, column=1, sticky=tk.W, padx=10)
            
            # Graph type dropdown
            ttk.Label(instruments_frame, text="Graph:", font=('Arial', 9)).grid(
                row=row_idx, column=2, sticky=tk.E, padx=5)
            
            graph_combo = ttk.Combobox(
                instruments_frame,
                textvariable=self.instrument_graph_types[inst_id],
                values=[label for _, label in graph_type_options],
                state='readonly',
                width=15
            )
            graph_combo.grid(row=row_idx, column=3, sticky=tk.W, padx=5)
            
            row_idx += 1
        
        instruments_frame.columnconfigure(1, weight=1)
        
        # Data aggregation (for power measurements)
        metric_frame = ttk.LabelFrame(tab, text="Data Aggregation (Power Rails)", padding="10")
        metric_frame.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=10, padx=(0, 5))
        
        ttk.Label(metric_frame, text="For power measurements, show:", 
                 font=('Arial', 10, 'bold')).grid(row=0, column=0, sticky=tk.W, pady=5)
        
        metric_types = [
            ("mean", "Mean (Average)"),
            ("min", "Minimum"),
            ("max", "Maximum"),
            ("all", "All (Min, Mean, Max)")
        ]
        
        for i, (value, label) in enumerate(metric_types):
            ttk.Radiobutton(metric_frame, text=label, variable=self.metric_type_var, value=value).grid(
                row=i+1, column=0, sticky=tk.W, padx=5, pady=2)
        
        # Data Source Selection
        source_frame = ttk.LabelFrame(tab, text="Data Source", padding="10")
        source_frame.grid(row=2, column=1, sticky=(tk.W, tk.E, tk.N), padx=5, pady=10)
        
        ttk.Label(source_frame, text="Select data source:", 
                 font=('Arial', 10, 'bold')).grid(row=0, column=0, sticky=tk.W, pady=5)
        
        ttk.Radiobutton(source_frame, text="results.json\n(Aggregated)", 
                       variable=self.data_source_var, value="results_json").grid(
            row=1, column=0, sticky=tk.W, padx=5, pady=5)
        
        ttk.Radiobutton(source_frame, text="TDMS\n(Time-series)", 
                       variable=self.data_source_var, value="tdms").grid(
            row=2, column=0, sticky=tk.W, padx=5, pady=5)
        
        # AI Trend Analysis Option
        ai_frame = ttk.LabelFrame(tab, text="AI Analysis (GENI)", padding="10")
        ai_frame.grid(row=2, column=2, sticky=(tk.W, tk.E, tk.N), pady=10, padx=(5, 0))
        
        ttk.Label(ai_frame, text="AI-powered analysis:", 
                 font=('Arial', 10, 'bold')).grid(row=0, column=0, sticky=tk.W, pady=5)
        
        ai_checkbox = ttk.Checkbutton(ai_frame, 
                                      text="Include GENI\nTrend Analysis",
                                      variable=self.include_trends_var)
        ai_checkbox.grid(row=1, column=0, sticky=tk.W, padx=5, pady=5)
        
        ttk.Label(ai_frame, text="Analyzes trends and\nprovides recommendations", 
                 foreground="gray", font=('Arial', 8)).grid(row=2, column=0, sticky=tk.W, padx=5, pady=2)
        
        # HSDES Sighting Query Option
        hsdes_frame = ttk.LabelFrame(tab, text="HSDES Query", padding="10")
        hsdes_frame.grid(row=3, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=10)
        
        hsdes_checkbox = ttk.Checkbutton(hsdes_frame, 
                                         text="Query HSDES for related power sightings",
                                         variable=self.query_hsdes_var)
        hsdes_checkbox.grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
        
        ttk.Label(hsdes_frame, text="Project Release (optional):", 
                 font=('Arial', 9)).grid(row=0, column=1, sticky=tk.W, padx=20)
        ttk.Entry(hsdes_frame, textvariable=self.project_release_var, width=15).grid(
            row=0, column=2, sticky=tk.W, padx=5)
        ttk.Label(hsdes_frame, text="e.g., NVL-2026.1", 
                 foreground="gray", font=('Arial', 7)).grid(row=0, column=3, sticky=tk.W, padx=5)
        
        # Output file selection
        output_frame = ttk.Frame(tab)
        output_frame.grid(row=4, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=10)
        
        ttk.Label(output_frame, text="Output HTML File:", 
                 font=('Arial', 10, 'bold')).grid(row=0, column=0, sticky=tk.W, pady=5)
        ttk.Entry(output_frame, textvariable=self.output_path_var, width=50).grid(row=1, column=0, sticky=(tk.W, tk.E))
        ttk.Button(output_frame, text="Browse...", command=self.browse_output).grid(row=1, column=1, padx=5)
        
        output_frame.columnconfigure(0, weight=1)
        
        # Export Section
        export_frame = ttk.LabelFrame(tab, text="Export Options", padding="10")
        export_frame.grid(row=5, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=10)
        
        ttk.Label(export_frame, text="Export selected data:", 
                 font=('Arial', 9)).grid(row=0, column=0, sticky=tk.W, pady=2)
        
        # Excel Export Button
        excel_btn = ttk.Button(export_frame, text="Export to Excel (.xlsx)", 
                              command=self.export_to_excel)
        excel_btn.grid(row=0, column=1, padx=5)
        
        # HTML Export Button
        html_btn = ttk.Button(export_frame, text="Export to HTML (Static)", 
                             command=self.export_to_html)
        html_btn.grid(row=0, column=2, padx=5)
        
        # Summary section
        ttk.Label(tab, text="Selection Summary:", 
                 font=('Arial', 10, 'bold')).grid(row=6, column=0, columnspan=3, pady=5, sticky=tk.W)
        
        summary_frame = ttk.Frame(tab, relief=tk.RIDGE, borderwidth=2, padding="5")
        summary_frame.grid(row=7, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=5)
        
        self.summary_text = tk.Text(summary_frame, height=8, width=90, wrap=tk.WORD)
        self.summary_text.pack(fill=tk.BOTH, expand=True)
        
        # Generate Interactive Dashboard button
        generate_btn = ttk.Button(tab, text="Generate Multi-Instrument Dashboard", 
                                 command=self.generate_dashboard,
                                 style='Accent.TButton')
        generate_btn.grid(row=8, column=0, columnspan=3, pady=15)
        
        # Progress section
        self.progress_var = tk.StringVar(value="")
        progress_label = ttk.Label(tab, textvariable=self.progress_var, foreground="blue")
        progress_label.grid(row=9, column=0, columnspan=3, pady=5)
        
        self.progress_bar = ttk.Progressbar(tab, mode='indeterminate')
        self.progress_bar.grid(row=10, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=5)
        
        tab.columnconfigure(0, weight=1)
        tab.columnconfigure(1, weight=1)
        tab.columnconfigure(2, weight=1)
    
    # [Previous helper methods remain the same: browse_folder, browse_output, scan_folders, etc.]
    def browse_folder(self):
        """Browse for results folder."""
        folder = filedialog.askdirectory(title="Select Hopper Results Folder")
        if folder:
            self.folder_path_var.set(folder)
    
    def browse_output(self):
        """Browse for output HTML file."""
        file = filedialog.asksaveasfilename(
            title="Save Dashboard As",
            defaultextension=".html",
            filetypes=[("HTML files", "*.html"), ("All files", "*.*")]
        )
        if file:
            self.output_path_var.set(file)
    
    def scan_folders(self):
        """Recursively scan the selected folder for results.json files."""
        folder_path = self.folder_path_var.get()
        if not folder_path:
            messagebox.showerror("Error", "Please select a folder first")
            return
        
        parent_folder = Path(folder_path)
        if not parent_folder.exists():
            messagebox.showerror("Error", f"Folder does not exist: {folder_path}")
            return
        
        self.status_var.set("Scanning folders recursively...")
        self.results_folders = []
        self.all_power_rails = set()
        self.all_socwatch_metrics = set()
        self.all_workloads = set()
        
        self.folders_listbox.delete(0, tk.END)
        
        # Recursively search for *-results.json files
        results_files_found = list(parent_folder.rglob("*-results.json"))
        
        logger.info(f"Found {len(results_files_found)} results.json files")
        
        for results_file in results_files_found:
            folder = results_file.parent
            
            # Skip if we already processed this folder
            if folder in self.results_folders:
                continue
            
            self.results_folders.append(folder)
            
            # Create workload name from relative path to make it unique
            try:
                rel_path = folder.relative_to(parent_folder)
                workload_name = str(rel_path).replace('\\', '/')  # Use / for consistency
            except ValueError:
                workload_name = folder.name
            
            self.all_workloads.add(workload_name)
            self.folder_name_map[workload_name] = folder  # Store mapping
            
            # Show with indent level to indicate nesting
            indent = "  " * (len(folder.relative_to(parent_folder).parents) - 1)
            self.folders_listbox.insert(tk.END, f"✓ {indent}{workload_name}")
            
            try:
                with open(results_file, 'r', encoding='utf-8') as f:
                    results_json = json.load(f)
                
                # Extract power rails and SocWatch metrics
                hopper = results_json.get('hopper', {})
                for subtest in hopper.get('subtests', []):
                    for group in subtest.get('result_groups', []):
                        if group.get('name') == 'flexlogger':
                            for result in group.get('results', []):
                                rail_name = result.get('name', '')
                                if rail_name.startswith('I_'):
                                    self.all_power_rails.add(rail_name)
                        
                        elif group.get('name') == 'socwatch':
                            for result in group.get('results', []):
                                metric_name = result.get('name', '')
                                if metric_name:
                                    self.all_socwatch_metrics.add(metric_name)
                
                # Also extract from 'power' section (DAQ data)
                power = results_json.get('power', {})
                for rail_name in power.keys():
                    if rail_name.startswith('I_') or rail_name.startswith('P_') or rail_name.startswith('V_'):
                        self.all_power_rails.add(rail_name)
            
            except Exception as e:
                logger.warning(f"Error parsing {results_file}: {e}")
        
        self.populate_rail_checkboxes()
        self.populate_socwatch_checkboxes()
        self.populate_workload_checkboxes()
        
        self.status_var.set(f"Found {len(self.results_folders)} results folders (searched recursively)")
        messagebox.showinfo("Scan Complete", 
                           f"Found (recursive search):\n{len(self.results_folders)} workload runs\n"
                           f"{len(self.all_power_rails)} power rails\n"
                           f"{len(self.all_socwatch_metrics)} SocWatch metrics\n"
                           f"{len(self.all_workloads)} workloads")
    
    def populate_rail_checkboxes(self):
        """Populate power rail checkboxes."""
        for widget in self.rails_checkbox_frame.winfo_children():
            widget.destroy()
        self.rail_checkboxes = {}
        
        sorted_rails = sorted(self.all_power_rails)
        for i, rail in enumerate(sorted_rails):
            var = tk.BooleanVar(value=True)
            cb = ttk.Checkbutton(self.rails_checkbox_frame, text=rail, variable=var)
            cb.grid(row=i, column=0, sticky=tk.W, padx=5, pady=2)
            self.rail_checkboxes[rail] = {'var': var, 'widget': cb}
    
    def populate_socwatch_checkboxes(self):
        """Populate SocWatch metric checkboxes."""
        for widget in self.socwatch_checkbox_frame.winfo_children():
            widget.destroy()
        self.socwatch_checkboxes = {}
        
        sorted_metrics = sorted(self.all_socwatch_metrics)
        for i, metric in enumerate(sorted_metrics):
            var = tk.BooleanVar(value=True)
            cb = ttk.Checkbutton(self.socwatch_checkbox_frame, text=metric, variable=var)
            cb.grid(row=i, column=0, sticky=tk.W, padx=5, pady=2)
            self.socwatch_checkboxes[metric] = {'var': var, 'widget': cb}
    
    def populate_workload_checkboxes(self):
        """Populate workload checkboxes."""
        for widget in self.workload_checkbox_frame.winfo_children():
            widget.destroy()
        self.workload_checkboxes = {}
        
        sorted_workloads = sorted(self.all_workloads)
        for i, workload in enumerate(sorted_workloads):
            var = tk.BooleanVar(value=True)
            cb = ttk.Checkbutton(self.workload_checkbox_frame, text=workload, variable=var)
            cb.grid(row=i, column=0, sticky=tk.W, padx=5, pady=2)
            self.workload_checkboxes[workload] = {'var': var, 'widget': cb}
    
    def filter_rails(self):
        """Filter power rails by keyword."""
        keyword = self.rail_filter_var.get().lower()
        for rail, checkbox_data in self.rail_checkboxes.items():
            if keyword in rail.lower():
                checkbox_data['widget'].grid()
            else:
                checkbox_data['widget'].grid_remove()
    
    def filter_socwatch(self):
        """Filter SocWatch metrics by keyword."""
        keyword = self.socwatch_filter_var.get().lower()
        for metric, checkbox_data in self.socwatch_checkboxes.items():
            if keyword in metric.lower():
                checkbox_data['widget'].grid()
            else:
                checkbox_data['widget'].grid_remove()
    
    def filter_workloads(self):
        """Filter workloads by keyword."""
        keyword = self.workload_filter_var.get().lower()
        for workload, checkbox_data in self.workload_checkboxes.items():
            if keyword in workload.lower():
                checkbox_data['widget'].grid()
            else:
                checkbox_data['widget'].grid_remove()
    
    def clear_rail_filter(self):
        """Clear rail filter and show all."""
        self.rail_filter_var.set("")
        self.filter_rails()
    
    def clear_socwatch_filter(self):
        """Clear SocWatch filter and show all."""
        self.socwatch_filter_var.set("")
        self.filter_socwatch()
    
    def clear_workload_filter(self):
        """Clear workload filter and show all."""
        self.workload_filter_var.set("")
        self.filter_workloads()
    
    def select_all_rails(self):
        """Select all visible power rails."""
        for rail, checkbox_data in self.rail_checkboxes.items():
            if checkbox_data['widget'].winfo_ismapped():
                checkbox_data['var'].set(True)
    
    def deselect_all_rails(self):
        """Deselect all power rails."""
        for checkbox_data in self.rail_checkboxes.values():
            checkbox_data['var'].set(False)
    
    def select_all_socwatch(self):
        """Select all visible SocWatch metrics."""
        for metric, checkbox_data in self.socwatch_checkboxes.items():
            if checkbox_data['widget'].winfo_ismapped():
                checkbox_data['var'].set(True)
    
    def deselect_all_socwatch(self):
        """Deselect all SocWatch metrics."""
        for checkbox_data in self.socwatch_checkboxes.values():
            checkbox_data['var'].set(False)
    
    def select_all_workloads(self):
        """Select all visible workloads."""
        for workload, checkbox_data in self.workload_checkboxes.items():
            if checkbox_data['widget'].winfo_ismapped():
                checkbox_data['var'].set(True)
    
    def deselect_all_workloads(self):
        """Deselect all workloads."""
        for checkbox_data in self.workload_checkboxes.values():
            checkbox_data['var'].set(False)
    
    def get_selected_rails(self) -> List[str]:
        """Get list of selected power rails."""
        return [rail for rail, data in self.rail_checkboxes.items() if data['var'].get()]
    
    def get_selected_socwatch(self) -> List[str]:
        """Get list of selected SocWatch metrics."""
        return [metric for metric, data in self.socwatch_checkboxes.items() if data['var'].get()]
    
    def get_selected_workloads(self) -> List[str]:
        """Get list of selected workloads."""
        return [workload for workload, data in self.workload_checkboxes.items() if data['var'].get()]
    
    def update_summary(self):
        """Update the selection summary."""
        selected_rails = self.get_selected_rails()
        selected_socwatch = self.get_selected_socwatch()
        selected_workloads = self.get_selected_workloads()
        
        graph_type_names = {
            "bar": "Bar Chart",
            "grouped_bar": "Grouped Bar Chart",
            "line": "Line Graph",
            "scatter": "Scatter Plot",
            "heatmap": "Heatmap/Matrix",
            "box": "Box Plot"
        }
        
        metric_type_names = {
            "mean": "Mean (Average)",
            "min": "Minimum",
            "max": "Maximum",
            "all": "All (Min, Mean, Max)"
        }
        
        # V3.4: Check if multi-instrument mode is enabled
        enabled_instruments = {
            inst: self.instrument_enabled[inst].get()
            for inst in ['daq', 'socwatch', 'perftracer', 'powertrace']
        }
        
        any_instrument_enabled = any(enabled_instruments.values())
        
        if any_instrument_enabled:
            # Multi-instrument mode summary
            instrument_lines = []
            instrument_names = {
                'daq': 'DAQ',
                'socwatch': 'SocWatch',
                'perftracer': 'PerfTracer',
                'powertrace': 'PowerTrace'
            }
            
            for inst_id in ['daq', 'socwatch', 'perftracer', 'powertrace']:
                if enabled_instruments[inst_id]:
                    graph_type = self.instrument_graph_types[inst_id].get()
                    graph_name = graph_type_names.get(graph_type, graph_type.title())
                    instrument_lines.append(f"   ✓ {instrument_names[inst_id]}: {graph_name}")
            
            instruments_summary = "\n".join(instrument_lines) if instrument_lines else "   (none selected)"
            
            summary = f"""
🔧 Mode: Multi-Instrument Dashboard (v3.4)
📊 Instruments Enabled:
{instruments_summary}
📈 Data Type: {metric_type_names.get(self.metric_type_var.get(), 'Unknown')}
✓ Workloads: {len(selected_workloads)} selected
⚡ Power Rails: {len(selected_rails)} selected
📉 SocWatch: {len(selected_socwatch)} selected
🤖 GENI Analysis: {'Enabled' if self.include_trends_var.get() else 'Disabled'}
🔍 HSDES Query: {'Enabled' if self.query_hsdes_var.get() else 'Disabled'}
💾 Output: {self.output_path_var.get()}
"""
        else:
            # Legacy single-graph mode summary
            summary = f"""
🔧 Mode: Single Graph (Legacy)
📊 Graph Type: {graph_type_names.get(self.graph_type_var.get(), 'Unknown')}
📈 Data Type: {metric_type_names.get(self.metric_type_var.get(), 'Unknown')}
✓ Workloads: {len(selected_workloads)} selected
⚡ Power Rails: {len(selected_rails)} selected
📉 SocWatch: {len(selected_socwatch)} selected
🤖 GENI Analysis: {'Enabled' if self.include_trends_var.get() else 'Disabled'}
🔍 HSDES Query: {'Enabled' if self.query_hsdes_var.get() else 'Disabled'}
💾 Output: {self.output_path_var.get()}
"""
        self.summary_text.delete('1.0', tk.END)
        self.summary_text.insert('1.0', summary)
    
    def generate_dashboard(self):
        """Generate the dashboard in a separate thread."""
        self.update_summary()
        
        if not self.results_folders:
            messagebox.showerror("Error", "No results folders scanned. Please scan first.")
            return
        
        selected_workloads = self.get_selected_workloads()
        if not selected_workloads:
            messagebox.showerror("Error", "Please select at least one workload")
            return
        
        selected_rails = self.get_selected_rails()
        selected_socwatch = self.get_selected_socwatch()
        
        if not selected_rails and not selected_socwatch:
            messagebox.showerror("Error", "Please select at least one metric (power rail or SocWatch)")
            return
        
        self.progress_var.set("Generating dashboard...")
        self.progress_bar.start()
        
        thread = threading.Thread(target=self._generate_dashboard_thread)
        thread.start()
    
    def _generate_dashboard_thread(self):
        """Generate dashboard in background thread."""
        try:
            selected_rails = self.get_selected_rails()
            selected_socwatch = self.get_selected_socwatch()
            selected_workloads = self.get_selected_workloads()
            output_path = Path(self.output_path_var.get())
            graph_type = self.graph_type_var.get()
            metric_type = self.metric_type_var.get()
            data_source = self.data_source_var.get()  # V3: Get data source
            
            # V3: Parse based on data source
            if data_source == "tdms":
                self.root.after(0, lambda: self.status_var.set("Parsing TDMS files..."))
                df = self.parse_tdms_with_filters(selected_workloads, selected_rails, metric_type)
            else:
                self.root.after(0, lambda: self.status_var.set("Parsing results.json files..."))
                df = self.parse_results_with_filters(selected_workloads, selected_rails, selected_socwatch, metric_type)
            
            # V3.2: Check if GENI trend analysis is enabled
            geni_html = ""
            if self.include_trends_var.get():
                self.root.after(0, lambda: self.status_var.set("Running GENI trend analysis..."))
                geni_html = self.run_geni_analysis(df, selected_rails + selected_socwatch)
            
            # V3.3: Check if HSDES sighting query is enabled
            hsdes_html = ""
            if self.query_hsdes_var.get():
                self.root.after(0, lambda: self.status_var.set("Querying HSDES sightings..."))
                hsdes_html = self.run_hsdes_query(selected_workloads)
            
            # V3.4: Check if multi-instrument mode is enabled (any instrument checkbox enabled)
            enabled_instruments = {
                inst: self.instrument_enabled[inst].get()
                for inst in ['daq', 'socwatch', 'perftracer', 'powertrace']
            }
            
            any_instrument_enabled = any(enabled_instruments.values())
            
            if any_instrument_enabled:
                # V3.4: Multi-instrument dashboard mode
                self.root.after(0, lambda: self.status_var.set("Creating multi-instrument dashboard..."))
                
                graph_types = {
                    inst: self.instrument_graph_types[inst].get()
                    for inst in ['daq', 'socwatch', 'perftracer', 'powertrace']
                }
                
                self.create_multi_instrument_dashboard(
                    df, output_path, enabled_instruments, graph_types,
                    metric_type, selected_rails, selected_socwatch, selected_workloads,
                    geni_html, hsdes_html
                )
            else:
                # Legacy mode: Single graph type for all data
                self.root.after(0, lambda: self.status_var.set(f"Creating {graph_type} plot..."))
                self.create_comparison_dashboard(df, output_path, graph_type, metric_type,
                                                selected_rails, selected_socwatch, selected_workloads, geni_html, hsdes_html)
            
            self.root.after(0, lambda: self.progress_bar.stop())
            self.root.after(0, lambda: self.progress_var.set("Dashboard generated successfully!"))
            self.root.after(0, lambda: self.status_var.set(f"Dashboard saved to: {output_path}"))
            self.root.after(0, lambda: messagebox.showinfo("Success", 
                f"Dashboard generated successfully!\n\nSaved to:\n{output_path}"))
        
        except Exception as e:
            logger.error(f"Error generating dashboard: {e}", exc_info=True)
            self.root.after(0, lambda: self.progress_bar.stop())
            self.root.after(0, lambda: self.progress_var.set("Error generating dashboard"))
            self.root.after(0, lambda: messagebox.showerror("Error", f"Failed to generate dashboard:\n{str(e)}"))
    
    def parse_results_with_filters(self, selected_workloads, selected_rails, selected_socwatch, metric_type) -> pd.DataFrame:
        """Parse results.json with filters applied."""
        all_data = []
        
        for folder in self.results_folders:
            # Use the full workload name from folder_name_map
            workload_name = None
            for wl_name, wl_folder in self.folder_name_map.items():
                if wl_folder == folder:
                    workload_name = wl_name
                    break
            
            if not workload_name or workload_name not in selected_workloads:
                continue
            
            results_files = list(folder.glob("*-results.json"))
            if not results_files:
                continue
            
            try:
                with open(results_files[0], 'r', encoding='utf-8') as f:
                    results_json = json.load(f)
                
                # Extract from hopper subtests (flexlogger and socwatch)
                hopper = results_json.get('hopper', {})
                for subtest in hopper.get('subtests', []):
                    for group in subtest.get('result_groups', []):
                        group_name = group.get('name', '')
                        results = group.get('results', [])
                        
                        for result in results:
                            metric_name = result.get('name', 'unknown')
                            value = result.get('value', '')
                            unit = result.get('unit', '')
                            meas_type = result.get('type', '')
                            
                            # Filter by selection
                            if group_name == 'flexlogger' and metric_name not in selected_rails:
                                continue
                            if group_name == 'socwatch' and metric_name not in selected_socwatch:
                                continue
                            
                            # Filter by metric type
                            if metric_type != 'all' and group_name == 'flexlogger':
                                if metric_type == 'mean' and 'mean' not in meas_type.lower():
                                    continue
                                elif metric_type == 'min' and 'min' not in meas_type.lower():
                                    continue
                                elif metric_type == 'max' and 'max' not in meas_type.lower():
                                    continue
                            
                            all_data.append({
                                'workload': workload_name,
                                'instrument': group_name,
                                'metric': metric_name,
                                'type': meas_type,
                                'value': value,
                                'unit': unit
                            })
                
                # Also extract from 'power' section (DAQ data)
                power = results_json.get('power', {})
                for rail_name, rail_data in power.items():
                    # Filter by selection
                    if rail_name not in selected_rails:
                        continue
                    
                    # Extract default measurements
                    if isinstance(rail_data, dict) and 'default' in rail_data:
                        default_data = rail_data['default']
                        
                        # Add based on metric_type filter
                        if metric_type == 'all' or metric_type == 'mean':
                            mean_value = default_data.get('mean')
                            unit = default_data.get('unit', '')
                            if mean_value is not None:
                                all_data.append({
                                    'workload': workload_name,
                                    'instrument': 'daq',  # Mark as DAQ data
                                    'metric': rail_name,
                                    'type': 'Power mean',
                                    'value': mean_value,
                                    'unit': unit
                                })
                        
                        if metric_type == 'all' or metric_type == 'min':
                            min_value = default_data.get('min')
                            unit = default_data.get('unit', '')
                            if min_value is not None:
                                all_data.append({
                                    'workload': workload_name,
                                    'instrument': 'daq',
                                    'metric': rail_name,
                                    'type': 'Power min',
                                    'value': min_value,
                                    'unit': unit
                                })
                        
                        if metric_type == 'all' or metric_type == 'max':
                            max_value = default_data.get('max')
                            unit = default_data.get('unit', '')
                            if max_value is not None:
                                all_data.append({
                                    'workload': workload_name,
                                    'instrument': 'daq',
                                    'metric': rail_name,
                                    'type': 'Power max',
                                    'value': max_value,
                                    'unit': unit
                                })
            
            except Exception as e:
                logger.error(f"Error parsing {results_files[0]}: {e}")
        
        return pd.DataFrame(all_data)
    
    def parse_tdms_with_filters(self, selected_workloads, selected_rails, metric_type) -> pd.DataFrame:
        """Parse TDMS files with filters applied (V3 NEW)."""
        all_data = []
        
        for folder in self.results_folders:
            workload_name = folder.name  # V3: Keep full folder name
            
            if workload_name not in selected_workloads:
                continue
            
            # Find TDMS files
            tdms_files = list(folder.glob("*-power-*.tdms"))
            if not tdms_files:
                logger.warning(f"No TDMS files found in {folder}")
                continue
            
            try:
                tdms_file = TdmsFile.read(tdms_files[0])
                
                for group in tdms_file.groups():
                    for channel in group.channels():
                        rail_name = channel.name
                        
                        # Filter by selection
                        if rail_name not in selected_rails:
                            continue
                        
                        # Get channel data
                        channel_data = channel[:]
                        
                        # Calculate statistics
                        mean_val = np.mean(channel_data)
                        min_val = np.min(channel_data)
                        max_val = np.max(channel_data)
                        
                        # Add based on metric_type filter
                        if metric_type == 'all' or metric_type == 'mean':
                            all_data.append({
                                'workload': workload_name,
                                'instrument': 'flexlogger',
                                'metric': rail_name,
                                'type': 'Current mean',
                                'value': mean_val,
                                'unit': 'A'
                            })
                        
                        if metric_type == 'all' or metric_type == 'min':
                            all_data.append({
                                'workload': workload_name,
                                'instrument': 'flexlogger',
                                'metric': rail_name,
                                'type': 'Current min',
                                'value': min_val,
                                'unit': 'A'
                            })
                        
                        if metric_type == 'all' or metric_type == 'max':
                            all_data.append({
                                'workload': workload_name,
                                'instrument': 'flexlogger',
                                'metric': rail_name,
                                'type': 'Current max',
                                'value': max_val,
                                'unit': 'A'
                            })
            
            except Exception as e:
                logger.error(f"Error parsing TDMS file {tdms_files[0]}: {e}")
        
        return pd.DataFrame(all_data)
    
    def export_to_excel(self):
        """Export selected metrics to Excel (V3 NEW)."""
        try:
            # Get selections
            selected_rails = self.get_selected_rails()
            selected_socwatch = self.get_selected_socwatch()
            selected_workloads = self.get_selected_workloads()
            metric_type = self.metric_type_var.get()
            data_source = self.data_source_var.get()
            
            if not selected_workloads:
                messagebox.showerror("Error", "Please select at least one workload")
                return
            
            if not selected_rails and not selected_socwatch:
                messagebox.showerror("Error", "Please select at least one metric")
                return
            
            # Parse data
            self.status_var.set("Parsing data for Excel export...")
            if data_source == "tdms":
                df = self.parse_tdms_with_filters(selected_workloads, selected_rails, metric_type)
            else:
                df = self.parse_results_with_filters(selected_workloads, selected_rails, selected_socwatch, metric_type)
            
            if df.empty:
                messagebox.showerror("Error", 
                    "No data to export!\n\n"
                    "Possible reasons:\n"
                    "• No matching metrics found in selected workloads\n"
                    "• Results files may not contain the selected metrics\n"
                    "• Try selecting different metrics or workloads\n\n"
                    f"Debug info:\n"
                    f"Selected workloads: {len(selected_workloads)}\n"
                    f"Selected rails: {len(selected_rails)}\n"
                    f"Selected SocWatch: {len(selected_socwatch)}")
                return
            
            df['value_numeric'] = pd.to_numeric(df['value'], errors='coerce')
            df_numeric = df[df['value_numeric'].notna()].copy()
            
            # Ask for save location
            output_path = filedialog.asksaveasfilename(
                title="Export to Excel",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )
            
            if not output_path:
                return  # User cancelled
            
            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Power KPI Data"
            
            # Get unique metrics
            unique_metrics = sorted(df_numeric['metric'].unique())
            
            # Header row
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            
            ws.cell(1, 1, "Folder Name").fill = header_fill
            ws.cell(1, 1).font = header_font
            ws.cell(1, 1).alignment = Alignment(horizontal="left", vertical="center")
            
            for col_idx, metric in enumerate(unique_metrics, start=2):
                ws.cell(1, col_idx, metric).fill = header_fill
                ws.cell(1, col_idx).font = header_font
                ws.cell(1, col_idx).alignment = Alignment(horizontal="center", vertical="center")
            
            # Data rows
            row_idx = 2
            for workload in selected_workloads:
                # First column: Full folder name (V3 requirement)
                ws.cell(row_idx, 1, workload)
                ws.cell(row_idx, 1).alignment = Alignment(horizontal="left")
                
                wl_data = df_numeric[df_numeric['workload'] == workload]
                
                for col_idx, metric in enumerate(unique_metrics, start=2):
                    metric_data = wl_data[wl_data['metric'] == metric]
                    if not metric_data.empty:
                        # Use mean if multiple values
                        value = metric_data['value_numeric'].mean()
                        ws.cell(row_idx, col_idx, round(value, 6))
                        ws.cell(row_idx, col_idx).alignment = Alignment(horizontal="right")
                
                row_idx += 1
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save workbook
            wb.save(output_path)
            self.status_var.set(f"Excel exported successfully")
            messagebox.showinfo("Success", 
                f"Data exported successfully!\n\n"
                f"Rows: {row_idx - 1} workloads\n"
                f"Columns: {len(unique_metrics)} metrics\n\n"
                f"Saved to:\n{output_path}")
        
        except Exception as e:
            logger.error(f"Error exporting to Excel: {e}", exc_info=True)
            messagebox.showerror("Error", f"Failed to export to Excel:\n{str(e)}")
    
    def export_to_html(self):
        """Export dashboard as static HTML file (V3 NEW)."""
        try:
            # Get selections
            selected_rails = self.get_selected_rails()
            selected_socwatch = self.get_selected_socwatch()
            selected_workloads = self.get_selected_workloads()
            graph_type = self.graph_type_var.get()
            metric_type = self.metric_type_var.get()
            data_source = self.data_source_var.get()
            
            if not selected_workloads:
                messagebox.showerror("Error", "Please select at least one workload")
                return
            
            if not selected_rails and not selected_socwatch:
                messagebox.showerror("Error", "Please select at least one metric")
                return
            
            # Ask for save location
            output_path = filedialog.asksaveasfilename(
                title="Export Dashboard to HTML",
                defaultextension=".html",
                filetypes=[("HTML files", "*.html"), ("All files", "*.*")],
                initialfile=f"power_kpi_{graph_type}_export.html"
            )
            
            if not output_path:
                return  # User cancelled
            
            output_path = Path(output_path)
            
            # Parse data
            self.status_var.set("Parsing data for HTML export...")
            if data_source == "tdms":
                df = self.parse_tdms_with_filters(selected_workloads, selected_rails, metric_type)
            else:
                df = self.parse_results_with_filters(selected_workloads, selected_rails, selected_socwatch, metric_type)
            
            if df.empty:
                messagebox.showerror("Error", 
                    "No data to export!\n\n"
                    "Possible reasons:\n"
                    "• No matching metrics found in selected workloads\n"
                    "• Results files may not contain the selected metrics\n"
                    "• Try selecting different metrics or workloads\n\n"
                    f"Debug info:\n"
                    f"Selected workloads: {len(selected_workloads)}\n"
                    f"Selected rails: {len(selected_rails)}\n"
                    f"Selected SocWatch: {len(selected_socwatch)}")
                return
            
            # Create dashboard
            self.status_var.set(f"Creating visualization...")
            
            # V3.2: Check if GENI trend analysis is enabled
            geni_html = ""
            if self.include_trends_var.get():
                self.status_var.set("Running GENI trend analysis...")
                geni_html = self.run_geni_analysis(df, selected_rails + selected_socwatch)
            
            # V3.4: Route through multi-instrument path when any instrument is enabled
            enabled_instruments = {
                inst: self.instrument_enabled[inst].get()
                for inst in ['daq', 'socwatch', 'perftracer', 'powertrace']
            }
            any_instrument_enabled = any(enabled_instruments.values())

            if any_instrument_enabled:
                self.status_var.set("Creating per-instrument HTML export...")
                graph_types = {
                    inst: self.instrument_graph_types[inst].get()
                    for inst in ['daq', 'socwatch', 'perftracer', 'powertrace']
                }
                self.create_multi_instrument_dashboard(
                    df, output_path, enabled_instruments, graph_types,
                    metric_type, selected_rails, selected_socwatch, selected_workloads,
                    geni_html
                )
            else:
                # Legacy single-graph mode
                graph_type = self.graph_type_var.get()
                self.create_comparison_dashboard(df, output_path, graph_type, metric_type,
                                                selected_rails, selected_socwatch, selected_workloads, geni_html)
            
            self.status_var.set(f"HTML exported successfully")
            messagebox.showinfo("Success", 
                f"Dashboard exported successfully!\n\n"
                f"Graph Type: {graph_type}\n"
                f"Workloads: {len(selected_workloads)}\n"
                f"Metrics: {len(selected_rails) + len(selected_socwatch)}\n\n"
                f"Saved to:\n{output_path}")
        
        except Exception as e:
            logger.error(f"Error exporting to HTML: {e}", exc_info=True)
            messagebox.showerror("Error", f"Failed to export to HTML:\n{str(e)}")
    
    def run_geni_analysis(self, df: pd.DataFrame, selected_metrics: List[str]) -> str:
        """
        Prepare GENI trend analysis request and save to file for agent processing.
        
        This method prepares the data summary and saves it to a file that the
        PowerKPI_Validator agent can process via GENI MCP.
        
        Args:
            df: DataFrame with workload data (columns: workload, instrument, metric, type, value, unit)
            selected_metrics: List of selected metric names
            
        Returns:
            str: HTML with instructions for agent processing
        """
        try:
            logger.info("Preparing GENI trend analysis request...")
            
            # Ensure we have numeric values for analysis
            # The df already has 'value_numeric' column if numeric conversion was done
            if 'value_numeric' not in df.columns:
                df['value_numeric'] = pd.to_numeric(df['value'], errors='coerce')
            
            # Filter to only numeric data
            df_numeric = df[df['value_numeric'].notna()].copy()
            
            if df_numeric.empty:
                logger.warning("No numeric data available for GENI analysis")
                return """
<div class="geni-error" style="margin: 20px 0; padding: 20px; background-color: #fff3cd; border-left: 5px solid #ffc107; border-radius: 5px;">
    <h2 style="color: #856404; margin-top: 0;">⚠️ GENI Trend Analysis - No Data</h2>
    <p style="color: #856404;">No numeric data available for analysis. Please select workloads with numerical metrics.</p>
</div>
"""
            
            # Prepare data summary
            summary_lines = []
            summary_lines.append("=== Power KPI Workload Test Results ===\n")
            
            # Add metadata
            summary_lines.append(f"Total Workloads Analyzed: {len(df_numeric['workload'].unique())}")
            summary_lines.append(f"Total Metrics Selected: {len(selected_metrics)}")
            summary_lines.append(f"Data Points: {len(df_numeric)} measurements\n")
            
            # Categorize metrics
            power_rails = [m for m in selected_metrics if m.startswith('I_') or 'POWER' in m.upper() or 'CURRENT' in m.upper()]
            socwatch_metrics = [m for m in selected_metrics if m not in power_rails]
            
            if power_rails:
                summary_lines.append(f"Power Rails Selected ({len(power_rails)}): {', '.join(power_rails[:5])}")
                if len(power_rails) > 5:
                    summary_lines.append(f"  ... and {len(power_rails) - 5} more")
            
            if socwatch_metrics:
                summary_lines.append(f"SocWatch Metrics Selected ({len(socwatch_metrics)}): {', '.join(socwatch_metrics[:5])}")
                if len(socwatch_metrics) > 5:
                    summary_lines.append(f"  ... and {len(socwatch_metrics) - 5} more")
            
            summary_lines.append("\n" + "="*60 + "\n")
            
            # Group by workload and metric, aggregate statistics
            for workload in df_numeric['workload'].unique():
                workload_data = df_numeric[df_numeric['workload'] == workload]
                summary_lines.append(f"\n## WORKLOAD: {workload}")
                summary_lines.append(f"   ({len(workload_data)} data points)")
                
                # Separate power rails and SocWatch metrics
                if power_rails:
                    summary_lines.append("\n   ### Power Rails:")
                    for metric in power_rails:
                        metric_data = workload_data[workload_data['metric'] == metric]
                        
                        if not metric_data.empty:
                            values = metric_data['value_numeric']
                            unit = metric_data['unit'].iloc[0] if 'unit' in metric_data.columns and not metric_data['unit'].empty else ''
                            
                            mean_val = values.mean()
                            min_val = values.min()
                            max_val = values.max()
                            std_val = values.std()
                            
                            unit_str = f" {unit}" if unit else ""
                            summary_lines.append(
                                f"     - {metric}: Mean={mean_val:.4f}{unit_str}, "
                                f"Min={min_val:.4f}{unit_str}, Max={max_val:.4f}{unit_str}, Std={std_val:.4f}{unit_str}"
                            )
                
                if socwatch_metrics:
                    summary_lines.append("\n   ### SocWatch Metrics:")
                    for metric in socwatch_metrics:
                        metric_data = workload_data[workload_data['metric'] == metric]
                        
                        if not metric_data.empty:
                            values = metric_data['value_numeric']
                            unit = metric_data['unit'].iloc[0] if 'unit' in metric_data.columns and not metric_data['unit'].empty else ''
                            
                            mean_val = values.mean()
                            min_val = values.min()
                            max_val = values.max()
                            std_val = values.std()
                            
                            unit_str = f" {unit}" if unit else ""
                            summary_lines.append(
                                f"     - {metric}: Mean={mean_val:.4f}{unit_str}, "
                                f"Min={min_val:.4f}{unit_str}, Max={max_val:.4f}{unit_str}, Std={std_val:.4f}{unit_str}"
                            )
            
            # Add comparison section
            summary_lines.append("\n\n" + "="*60)
            summary_lines.append("\n=== CROSS-WORKLOAD COMPARISON ===\n")
            
            if power_rails:
                summary_lines.append("### Power Rails Comparison:")
                for metric in power_rails:
                    metric_data = df_numeric[df_numeric['metric'] == metric]
                    
                    if not metric_data.empty:
                        unit = metric_data['unit'].iloc[0] if 'unit' in metric_data.columns and not metric_data['unit'].empty else ''
                        unit_str = f" {unit}" if unit else ""
                        
                        summary_lines.append(f"\n## {metric}")
                        
                        # Calculate for each workload and sort
                        workload_means = []
                        for workload in df_numeric['workload'].unique():
                            workload_metric_data = metric_data[metric_data['workload'] == workload]
                            if not workload_metric_data.empty:
                                mean_val = workload_metric_data['value_numeric'].mean()
                                workload_means.append((workload, mean_val))
                        
                        # Sort by mean value (ascending for power = better efficiency)
                        workload_means.sort(key=lambda x: x[1])
                        
                        for rank, (workload, mean_val) in enumerate(workload_means, 1):
                            rank_str = "🥇 BEST" if rank == 1 else ("🥈" if rank == 2 else ("🥉" if rank == 3 else f"  #{rank}"))
                            summary_lines.append(f"  {rank_str} {workload}: {mean_val:.4f}{unit_str}")
            
            if socwatch_metrics:
                summary_lines.append("\n### SocWatch Metrics Comparison:")
                for metric in socwatch_metrics:
                    metric_data = df_numeric[df_numeric['metric'] == metric]
                    
                    if not metric_data.empty:
                        unit = metric_data['unit'].iloc[0] if 'unit' in metric_data.columns and not metric_data['unit'].empty else ''
                        unit_str = f" {unit}" if unit else ""
                        
                        summary_lines.append(f"\n## {metric}")
                        for workload in df_numeric['workload'].unique():
                            workload_metric_data = metric_data[metric_data['workload'] == workload]
                            if not workload_metric_data.empty:
                                mean_val = workload_metric_data['value_numeric'].mean()
                                summary_lines.append(f"  - {workload}: {mean_val:.4f}{unit_str}")
            
            data_summary = "\n".join(summary_lines)
            
            # Save to file for agent processing
            output_path = Path(self.output_path_var.get())
            geni_request_file = output_path.parent / f"{output_path.stem}_geni_request.txt"
            
            with open(geni_request_file, 'w', encoding='utf-8') as f:
                f.write("=== GENI Trend Analysis Request ===\n\n")
                f.write("INSTRUCTIONS FOR POWERKI_VALIDATOR AGENT:\n")
                f.write("Please query GENI Focus Mode 12 (VE Wiki) or Focus Mode 5 (Debug Assistant)\n")
                f.write("with the following question:\n\n")
                f.write("---\n\n")
                f.write(f"""I have Power KPI workload test results from Intel platform validation. 
Please provide a comprehensive trend analysis with the following structure:

## 1. EXECUTIVE SUMMARY
Provide a high-level overview of key findings across all workloads and metrics.

## 2. PER-WORKLOAD ANALYSIS
For EACH workload, provide:
- **Summary**: Brief description of workload behavior
- **Trend Overview**: Analysis of each selected parameter (power rails and SocWatch metrics)
  - Identify if values are stable, increasing, decreasing, or volatile
  - Compare against expected ranges for this workload type
- **Key Observations**: Notable patterns, anomalies, or concerns
- **Health Assessment**: Is this workload performing as expected?

## 3. CROSS-WORKLOAD COMPARISON
Compare the selected parameters across all workloads:
- **Power Efficiency Ranking**: Best to worst for each power rail
- **SocWatch Metrics Comparison**: How do C-states, frequencies, and residencies compare?
- **Relative Performance**: Which workload is most/least efficient?
- **Delta Analysis**: Highlight significant differences (>10% variance)

## 4. ANOMALY DETECTION
Identify and explain:
- Outliers in power measurements
- Unexpected SocWatch metric values
- Inconsistencies between similar workloads
- Values outside typical validation ranges

## 5. ACTIONABLE RECOMMENDATIONS
Provide specific next steps:
- Debug actions for anomalies
- Configuration changes to try
- Additional metrics to investigate
- BIOS knobs or firmware to check

Here is the test data:

{data_summary}

Please format your response with clear headings, bullet points, and specific numerical references.
Focus on actionable insights that validation engineers can use immediately.
""")
            
            logger.info(f"GENI request saved to: {geni_request_file}")
            
            # Return HTML with instructions
            return f"""
<div class="geni-pending" style="margin: 20px 0; padding: 20px; background-color: #e3f2fd; border-left: 5px solid #2196f3; border-radius: 5px;">
    <h2 style="color: #1976d2; margin-top: 0;">
        <span style="font-size: 24px;">🤖</span> GENI Trend Analysis - Agent Processing Required
    </h2>
    <p style="color: #1565c0; font-size: 1.1em; font-weight: bold;">
        The AI trend analysis request has been prepared.
    </p>
    <div style="margin: 15px 0; padding: 15px; background-color: white; border-radius: 5px; border: 1px solid #90caf9;">
        <h3 style="color: #1976d2; margin-top: 0;">Next Steps:</h3>
        <ol style="color: #424242; line-height: 1.8;">
            <li>Ask the <strong>PowerKPI_Validator agent</strong> to process the GENI analysis</li>
            <li>Provide this file path: <code style="background-color: #f5f5f5; padding: 2px 6px; border-radius: 3px;">{geni_request_file}</code></li>
            <li>The agent will query GENI via MCP and inject the insights into this dashboard</li>
        </ol>
    </div>
    <div style="margin: 15px 0; padding: 15px; background-color: #fff9c4; border-radius: 5px; border: 1px solid #fbc02d;">
        <h3 style="color: #f57f17; margin-top: 0;">📝 Data Summary Preview:</h3>
        <pre style="white-space: pre-wrap; font-family: 'Consolas', 'Monaco', monospace; font-size: 0.9em; max-height: 300px; overflow-y: auto; color: #424242;">
{data_summary[:1000]}{'...' if len(data_summary) > 1000 else ''}
        </pre>
    </div>
    <div style="margin-top: 15px; padding-top: 15px; border-top: 1px solid #90caf9; font-size: 0.9em; color: #666;">
        <p><strong>Alternative:</strong> Run dashboard generation through the PowerKPI_Validator agent for automatic GENI integration.</p>
        <p><em>Example: "Please generate a Power KPI dashboard with GENI trend analysis for IDON and CMS workloads"</em></p>
    </div>
</div>
"""
                
        except Exception as e:
            logger.error(f"GENI analysis preparation error: {e}", exc_info=True)
            return f"""
<div class="geni-error" style="margin: 20px 0; padding: 20px; background-color: #fff3cd; border-left: 5px solid #ffc107; border-radius: 5px;">
    <h2 style="color: #856404; margin-top: 0;">
        <span style="font-size: 24px;">⚠️</span> GENI Trend Analysis Preparation Error
    </h2>
    <p style="color: #856404;">
        {str(e)}
    </p>
</div>
"""
    
    def run_hsdes_query(self, selected_workloads: List[str]) -> str:
        """
        Query HSDES sightings for power debugging based on results.json data.
        
        Args:
            selected_workloads: List of selected workload folder names
            
        Returns:
            str: HTML with HSDES sighting query results
        """
        try:
            if not HSDES_AVAILABLE:
                return """
<div class="hsdes-error" style="margin: 20px 0; padding: 20px; background-color: #fff3cd; border-left: 5px solid #ffc107; border-radius: 5px;">
    <h2 style="color: #856404; margin-top: 0;">⚠️ HSDES Integration Unavailable</h2>
    <p style="color: #856404;">HSDES power debugger requires <code>pysvtools</code> package. Please install it to enable this feature.</p>
</div>
"""
            
            logger.info("Running HSDES sighting queries for power debugging...")
            
            debugger = HSDESPowerDebugger()
            project_release = self.project_release_var.get() if self.project_release_var.get() else None
            
            all_reports = []
            
            # Query for each selected workload
            for workload_name in selected_workloads:
                folder = self.folder_name_map.get(workload_name)
                if not folder:
                    continue
                
                results_files = list(folder.glob("*-results.json"))
                if not results_files:
                    continue
                
                results_file = results_files[0]
                
                try:
                    logger.info(f"Generating HSDES report for: {workload_name}")
                    report = debugger.generate_debug_report(str(results_file), project_release)
                    all_reports.append(report)
                except Exception as e:
                    logger.warning(f"Failed to query HSDES for {workload_name}: {e}")
                    continue
            
            if not all_reports:
                return """
<div class="hsdes-error" style="margin: 20px 0; padding: 20px; background-color: #fff3cd; border-left: 5px solid #ffc107; border-radius: 5px;">
    <h2 style="color: #856404; margin-top: 0;">⚠️ HSDES Query - No Results</h2>
    <p style="color: #856404;">No HSDES sightings found for the selected workloads. Try adjusting project release filter or check your HSDES credentials.</p>
</div>
"""
            
            # Generate HTML report
            html_parts = []
            html_parts.append("""
<div class="hsdes-results" style="margin: 20px 0; padding: 20px; background-color: #f3e5f5; border-left: 5px solid #9c27b0; border-radius: 5px;">
    <h2 style="color: #6a1b9a; margin-top: 0;">
        <span style="font-size: 24px;">🔍</span> HSDES Power Debugging Results
    </h2>
    <p style="color: #4a148c; font-size: 1.1em;">
        Found related sightings for power and Package C-state issues.
    </p>
""")
            
            for i, report in enumerate(all_reports, 1):
                workload = report.get('workload', 'Unknown')
                context = report.get('debug_context', {})
                sightings = report.get('sightings', [])
                keywords = report.get('search_keywords', [])
                
                html_parts.append(f"""
    <div style="margin: 15px 0; padding: 15px; background-color: white; border-radius: 5px; border: 1px solid #ce93d8;">
        <h3 style="color: #6a1b9a; margin-top: 0;">
            <span style="font-size: 18px;">📊</span> Workload {i}: {workload}
        </h3>
        
        <div style="margin: 10px 0;">
            <h4 style="color: #8e24aa;">Debug Context:</h4>
            <ul style="color: #424242; line-height: 1.6;">
                <li><strong>Power Rails Analyzed:</strong> {len(context.get('high_power_rails', []))}</li>
                <li><strong>SocWatch C-States:</strong> {len(context.get('socwatch_metrics', {}))}</li>
                <li><strong>Bad Residency Issues:</strong> {len(context.get('bad_residency', []))}</li>
            </ul>
        </div>
""")
                
                # Show bad residency issues
                if context.get('bad_residency'):
                    html_parts.append("""
        <div style="margin: 10px 0; padding: 10px; background-color: #ffebee; border-left: 3px solid #f44336; border-radius: 3px;">
            <h4 style="color: #c62828; margin-top: 0;">⚠️ Package C-State Issues Detected:</h4>
            <ul style="color: #d32f2f;">
""")
                    for issue in context['bad_residency']:
                        html_parts.append(f"                <li>{issue['issue']}</li>\n")
                    html_parts.append("            </ul>\n        </div>\n")
                
                # Show top 3 high power rails
                high_power = sorted(context.get('high_power_rails', []), key=lambda x: x['mean'], reverse=True)[:3]
                if high_power:
                    html_parts.append("""
        <div style="margin: 10px 0; padding: 10px; background-color: #fff3e0; border-left: 3px solid #ff9800; border-radius: 3px;">
            <h4 style="color: #e65100; margin-top: 0;">⚡ Top High Power Rails:</h4>
            <ul style="color: #e65100;">
""")
                    for rail in high_power:
                        html_parts.append(f"                <li><strong>{rail['name']}</strong>: {rail['mean']:.4f} {rail['unit']}</li>\n")
                    html_parts.append("            </ul>\n        </div>\n")
                
                # Show HSDES sightings found
                html_parts.append(f"""
        <div style="margin: 10px 0;">
            <h4 style="color: #6a1b9a;">🔗 Related HSDES Sightings Found: {len(sightings)}</h4>
""")
                
                if sightings:
                    html_parts.append("            <table style='width: 100%; border-collapse: collapse; margin-top: 10px;'>\n")
                    html_parts.append("""
                <thead>
                    <tr style='background-color: #f3e5f5; border-bottom: 2px solid #9c27b0;'>
                        <th style='padding: 8px; text-align: left; color: #6a1b9a;'>ID</th>
                        <th style='padding: 8px; text-align: left; color: #6a1b9a;'>Title</th>
                        <th style='padding: 8px; text-align: left; color: #6a1b9a;'>Status</th>
                        <th style='padding: 8px; text-align: left; color: #6a1b9a;'>Owner</th>
                    </tr>
                </thead>
                <tbody>
""")
                    for sighting in sightings[:10]:  # Show top 10
                        sid = sighting.get('id', 'N/A')
                        title = sighting.get('title', 'N/A')
                        status = sighting.get('status', 'N/A')
                        owner = sighting.get('owner', 'N/A')
                        
                        html_parts.append(f"""
                    <tr style='border-bottom: 1px solid #e1bee7;'>
                        <td style='padding: 8px;'><a href='https://hsdes.intel.com/appstore/article/#{sid}' target='_blank' style='color: #7b1fa2; text-decoration: none;'>{sid}</a></td>
                        <td style='padding: 8px;'>{title[:80]}{'...' if len(title) > 80 else ''}</td>
                        <td style='padding: 8px;'>{status}</td>
                        <td style='padding: 8px;'>{owner}</td>
                    </tr>
""")
                    html_parts.append("                </tbody>\n            </table>\n")
                else:
                    html_parts.append("""
            <p style='color: #666; font-style: italic;'>No direct matches found. Consider broadening search criteria or checking different project release.</p>
""")
                
                html_parts.append("        </div>\n")
                
                # Show search keywords used
                if keywords:
                    html_parts.append(f"""
        <div style="margin: 10px 0; padding: 10px; background-color: #f5f5f5; border-radius: 3px;">
            <p style="font-size: 0.9em; color: #666; margin: 0;"><strong>Search Keywords Used:</strong> {', '.join(keywords[:10])}</p>
        </div>
""")
                
                html_parts.append("    </div>\n")
            
            # Add recommendations
            html_parts.append("""
    <div style="margin: 20px 0; padding: 15px; background-color: #e8f5e9; border-left: 3px solid #4caf50; border-radius: 5px;">
        <h3 style="color: #2e7d32; margin-top: 0;">💡 Next Steps:</h3>
        <ol style="color: #1b5e20; line-height: 1.8;">
            <li>Review related sightings for known issues and workarounds</li>
            <li>Check IFWI/BIOS version compatibility with reported issues</li>
            <li>Verify BIOS knobs match recommendations from sightings</li>
            <li>If no matches found, consider filing a new sighting with current data</li>
        </ol>
    </div>
</div>
""")
            
            return '\n'.join(html_parts)
                
        except Exception as e:
            logger.error(f"HSDES query error: {e}", exc_info=True)
            return f"""
<div class="hsdes-error" style="margin: 20px 0; padding: 20px; background-color: #fff3cd; border-left: 5px solid #ffc107; border-radius: 5px;">
    <h2 style="color: #856404; margin-top: 0;">
        <span style="font-size: 24px;">⚠️</span> HSDES Query Error
    </h2>
    <p style="color: #856404;">
        {str(e)}
    </p>
</div>
"""
    
    def create_comparison_dashboard(self, df: pd.DataFrame, output_path: Path, graph_type: str, metric_type: str,
                                    selected_rails: List[str], selected_socwatch: List[str], selected_workloads: List[str],
                                    geni_html: str = "", hsdes_html: str = ""):
        """Create comparison dashboard based on selected graph type (legacy single-graph mode)."""
        df['value_numeric'] = pd.to_numeric(df['value'], errors='coerce')
        df_numeric = df[df['value_numeric'].notna()].copy()

        # Determine dominant instrument for axis-label context
        instruments_present = df_numeric['instrument'].unique().tolist() if 'instrument' in df_numeric.columns else []
        if 'daq' in instruments_present or 'flexlogger' in instruments_present:
            instrument_ctx = 'daq'
        elif 'socwatch' in instruments_present:
            instrument_ctx = 'socwatch'
        elif instruments_present:
            instrument_ctx = instruments_present[0]
        else:
            instrument_ctx = 'daq'

        if graph_type == "bar":
            fig = self.create_bar_chart(df_numeric, selected_workloads, instrument=instrument_ctx)
        elif graph_type == "grouped_bar":
            fig = self.create_grouped_bar_chart(df_numeric, selected_workloads, instrument=instrument_ctx)
        elif graph_type == "line":
            fig = self.create_line_graph(df_numeric, selected_workloads, instrument=instrument_ctx)
        elif graph_type == "scatter":
            fig = self.create_scatter_plot(df_numeric, selected_workloads, instrument=instrument_ctx)
        elif graph_type == "heatmap":
            fig = self.create_heatmap(df_numeric, selected_workloads, instrument=instrument_ctx)
        elif graph_type == "box":
            fig = self.create_box_plot(df_numeric, selected_workloads, instrument=instrument_ctx)
        else:
            raise ValueError(f"Unknown graph type: {graph_type}")
        
        # Save with optional GENI trend analysis and/or HSDES sighting query
        if geni_html or hsdes_html:
            # Generate the plot as HTML string first
            plot_html = fig.to_html(
                config={
                    'displayModeBar': True,
                    'displaylogo': False,
                    'toImageButtonOptions': {
                        'format': 'png',
                        'filename': f'power_kpi_{graph_type}',
                        'height': 1000,
                        'width': 1600,
                        'scale': 2
                    }
                },
                include_plotlyjs='cdn'
            )
            
            # Build title based on enabled features
            title_parts = ["Power KPI Dashboard"]
            if geni_html:
                title_parts.append("AI-Powered Trend Analysis")
            if hsdes_html:
                title_parts.append("HSDES Power Debugging")
            title = " with " + " and ".join(title_parts[1:]) if len(title_parts) > 1 else title_parts[0]
            
            # Inject GENI and/or HSDES sections before the plot
            combined_html = f"""
<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <title>{title}</title>
    <style>
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            margin: 20px;
            background-color: #f5f5f5;
        }}
        .container {{
            max-width: 1800px;
            margin: 0 auto;
            background-color: white;
            padding: 20px;
            border-radius: 10px;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        }}
        h1 {{
            color: #0071c5;
            text-align: center;
        }}
    </style>
</head>
<body>
    <div class="container">
        <h1>{title}</h1>
        
        <!-- GENI Trend Analysis Section -->
        {geni_html if geni_html else ''}
        
        <!-- HSDES Sighting Query Section -->
        {hsdes_html if hsdes_html else ''}
        
        <!-- Power KPI Graph -->
        <div class="plot-section">
            {plot_html}
        </div>
    </div>
</body>
</html>
"""
            
            # Write combined HTML
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(combined_html)
        else:
            # Original behavior without GENI analysis
            fig.write_html(
                output_path,
                config={
                    'displayModeBar': True,
                    'displaylogo': False,
                    'toImageButtonOptions': {
                        'format': 'png',
                        'filename': f'power_kpi_{graph_type}',
                        'height': 1000,
                        'width': 1600,
                        'scale': 2
                    }
                }
            )
    
    def create_graph_by_type(self, df: pd.DataFrame, graph_type: str, workloads: List[str],
                             instrument: str = 'daq') -> go.Figure:
        """
        Create graph based on selected type, passing instrument context for correct
        axis labels and units.

        Args:
            df: DataFrame with power/socwatch data (must NOT mix instruments)
            graph_type: Graph type string (e.g., 'bar', 'line', 'scatter')
            workloads: List of selected workload names
            instrument: Instrument identifier ('daq', 'socwatch', 'perftracer', 'powertrace')

        Returns:
            go.Figure: Plotly figure object
        """
        # Normalise graph_type coming from the Combobox display labels
        label_to_key = {
            'Bar Chart': 'bar',
            'Grouped Bar': 'grouped_bar',
            'Line Graph': 'line',
            'Scatter Plot': 'scatter',
            'Heatmap': 'heatmap',
            'Box Plot': 'box',
        }
        graph_type = label_to_key.get(graph_type, graph_type)

        # Convert values to numeric
        df = df.copy()
        df['value_numeric'] = pd.to_numeric(df['value'], errors='coerce')
        df_numeric = df[df['value_numeric'].notna()].copy()

        graph_type_map = {
            'bar':         lambda d, w: self.create_bar_chart(d, w, instrument=instrument),
            'grouped_bar': lambda d, w: self.create_grouped_bar_chart(d, w, instrument=instrument),
            'line':        lambda d, w: self.create_line_graph(d, w, instrument=instrument),
            'scatter':     lambda d, w: self.create_scatter_plot(d, w, instrument=instrument),
            'heatmap':     lambda d, w: self.create_heatmap(d, w, instrument=instrument),
            'box':         lambda d, w: self.create_box_plot(d, w, instrument=instrument),
        }

        if graph_type not in graph_type_map:
            raise ValueError(f"Unknown graph type: {graph_type!r}")

        return graph_type_map[graph_type](df_numeric, workloads)
    
    def create_multi_instrument_dashboard(self, df: pd.DataFrame, output_path: Path, 
                                          enabled_instruments: Dict[str, bool], 
                                          graph_types: Dict[str, str],
                                          metric_type: str, selected_rails: List[str], 
                                          selected_socwatch: List[str], selected_workloads: List[str],
                                          geni_html: str = "", hsdes_html: str = ""):
        """
        Create dashboard with separate graphs per instrument.
        
        Args:
            df: DataFrame with all data
            output_path: Path to save HTML file
            enabled_instruments: Dict of instrument ID -> enabled bool
            graph_types: Dict of instrument ID -> graph type string
            metric_type: Metric type (mean/min/max)
            selected_rails: List of selected power rails
            selected_socwatch: List of selected SocWatch metrics
            selected_workloads: List of selected workload names
            geni_html: Optional GENI trend analysis HTML
            hsdes_html: Optional HSDES sighting query HTML
        """
        logger.info(f"Creating multi-instrument dashboard with enabled: {enabled_instruments}")
        
        # Separate data by instrument
        # Map internal instrument tags to logical instruments:
        # - daq/flexlogger -> DAQ instrument
        # - socwatch -> SocWatch instrument
        df_daq = df[df['instrument'].isin(['daq', 'flexlogger'])].copy()
        df_socwatch = df[df['instrument'] == 'socwatch'].copy()
        df_perftracer = df[df['instrument'] == 'perftracer'].copy()
        df_powertrace = df[df['instrument'] == 'powertrace'].copy()
        
        instrument_data = {
            'daq': df_daq,
            'socwatch': df_socwatch,
            'perftracer': df_perftracer,
            'powertrace': df_powertrace
        }
        
        instrument_names = {
            'daq': 'DAQ — Power Rails',
            'socwatch': 'SocWatch — C-States & Frequencies',
            'perftracer': 'PerfTracer — Performance Metrics',
            'powertrace': 'PowerTrace — Power Traces'
        }
        
        graphs_html = []
        
        # Generate graph(s) for each enabled instrument.
        # KEY RULE: after isolating each instrument's data, split further by unit
        # so that e.g. DAQ Watts, Amps and Volts each get their own chart, and
        # SocWatch % residency and MHz frequency are never on the same axis.
        for inst_id in ['daq', 'socwatch', 'perftracer', 'powertrace']:
            if not enabled_instruments.get(inst_id, False):
                logger.info(f"Skipping disabled instrument: {inst_id}")
                continue
            
            df_inst = instrument_data[inst_id]
            if df_inst.empty:
                logger.warning(f"No data for instrument: {inst_id}")
                continue
            
            graph_type = graph_types.get(inst_id, 'bar')

            # --- Split by unit: one graph per distinct unit ---
            unit_groups = self._split_by_unit(df_inst)
            logger.info(f"Instrument {inst_id}: {len(unit_groups)} unit group(s): "
                        f"{[u for u, _ in unit_groups]}")

            for unit_label, df_unit in unit_groups:
                if df_unit.empty:
                    continue

                # Build a human-readable sub-title suffix
                unit_suffix = f" [{unit_label}]" if unit_label and unit_label != 'unknown' else ""
                pane_name = f"{instrument_names[inst_id]}{unit_suffix}"
                # Unique pane id: instrument + sanitised unit (spaces → underscore)
                safe_unit = unit_label.replace(' ', '_').replace('/', '_').replace('%', 'pct')
                pane_id = f"{inst_id}_{safe_unit}" if unit_label else inst_id

                logger.info(f"  Creating {graph_type} graph for {inst_id} unit={unit_label!r} "
                            f"with {len(df_unit)} data points")

                try:
                    fig = self.create_graph_by_type(df_unit, graph_type, selected_workloads,
                                                    instrument=inst_id)

                    # Update title to include unit group
                    fig.update_layout(
                        title=f'{pane_name} — {graph_type.replace("_", " ").title()}',
                        height=600
                    )

                    graph_html = fig.to_html(
                        include_plotlyjs=False,
                        full_html=False,
                        div_id=f'plotly-graph-{pane_id}',
                        config={
                            'displayModeBar': True,
                            'displaylogo': False,
                            'toImageButtonOptions': {
                                'format': 'png',
                                'filename': f'power_kpi_{pane_id}_{graph_type}',
                                'height': 1000,
                                'width': 1600,
                                'scale': 2
                            }
                        }
                    )

                    graphs_html.append({
                        'instrument': inst_id,
                        'pane_id': pane_id,
                        'name': pane_name,
                        'unit': unit_label,
                        'graph_type': graph_type.replace('_', ' ').title(),
                        'html': graph_html
                    })

                except Exception as e:
                    logger.error(f"Failed to create graph for {inst_id} unit={unit_label!r}: {e}",
                                 exc_info=True)
                    continue
        
        if not graphs_html:
            raise ValueError("No graphs were generated. Please enable at least one instrument with data.")
        
        # Write combined HTML
        self.write_multi_instrument_html(output_path, graphs_html, geni_html, hsdes_html)
    
    def write_multi_instrument_html(self, output_path: Path, graphs_html: List[Dict[str, str]], 
                                     geni_html: str = "", hsdes_html: str = ""):
        """
        Write multi-instrument dashboard HTML file.

        Each (instrument, unit) pair gets its own independent Plotly pane so
        that incompatible units are never rendered on the same axis.

        graphs_html entries have keys:
            instrument, pane_id, name, unit, graph_type, html
        """
        # Build title
        title_parts = ["Multi-Instrument Power KPI Dashboard"]
        if geni_html:
            title_parts.append("AI-Powered Trend Analysis")
        if hsdes_html:
            title_parts.append("HSDES Power Debugging")
        title = " with " + " and ".join(title_parts[1:]) if len(title_parts) > 1 else title_parts[0]

        instrument_icons = {
            'daq': '⚡',
            'flexlogger': '⚡',
            'socwatch': '📊',
            'perftracer': '🔬',
            'powertrace': '🔋',
        }

        # Unit-to-colour mapping so visually related panes share a border colour
        # while still being completely separate charts.
        unit_colours = {}
        palette = ['#0071c5', '#e05c00', '#2e7d32', '#6a1b9a', '#c62828',
                   '#00838f', '#4e342e', '#37474f', '#ef6c00', '#1565c0']
        colour_idx = 0

        def _colour_for(instrument: str, unit: str) -> str:
            nonlocal colour_idx
            key = (instrument, unit)
            if key not in unit_colours:
                unit_colours[key] = palette[colour_idx % len(palette)]
                colour_idx += 1
            return unit_colours[key]

        # Build panes
        graphs_section = []
        for i, graph_info in enumerate(graphs_html, 1):
            icon = instrument_icons.get(graph_info['instrument'], '📈')
            colour = _colour_for(graph_info['instrument'], graph_info.get('unit', ''))
            unit_label = graph_info.get('unit', '')
            unit_badge = (f'<span style="background-color:#555;color:white;'
                          f'padding:4px 10px;border-radius:12px;font-size:0.8em;'
                          f'margin-left:8px;">{unit_label}</span>'
                          if unit_label and unit_label != 'unknown' else '')
            pane_id = graph_info.get('pane_id', graph_info['instrument'])

            graphs_section.append(f"""
        <!-- Pane {i}: {graph_info['name']} -->
        <div class="instrument-pane" id="pane-{pane_id}" style="
            margin: 40px 0;
            padding: 25px;
            background: linear-gradient(135deg, #ffffff 0%, #f8f9fa 100%);
            border-radius: 12px;
            border: 2px solid {colour};
            box-shadow: 0 4px 6px rgba(0,0,0,0.1);
        ">
            <div class="pane-header" style="
                display: flex;
                align-items: center;
                justify-content: space-between;
                margin-bottom: 20px;
                padding-bottom: 15px;
                border-bottom: 3px solid {colour};
            ">
                <h2 style="color: {colour}; margin: 0; font-size: 1.8em;">
                    {icon} {graph_info['name']}{unit_badge}
                </h2>
                <span style="
                    background-color: {colour};
                    color: white;
                    padding: 8px 16px;
                    border-radius: 20px;
                    font-size: 0.9em;
                    font-weight: bold;
                ">{graph_info['graph_type']}</span>
            </div>

            <!-- Independent Plotly Graph Container -->
            <div class="plotly-graph-container" style="
                background-color: white;
                padding: 20px;
                border-radius: 8px;
                border: 1px solid #e0e0e0;
                min-height: 600px;
            ">
                {graph_info['html']}
            </div>

            <div class="pane-footer" style="
                margin-top: 15px;
                padding-top: 10px;
                border-top: 1px solid #e0e0e0;
                text-align: right;
                color: #666;
                font-size: 0.85em;
            ">
                <em>Instrument: {graph_info['instrument'].upper()}
                {(" | Unit: " + unit_label) if unit_label and unit_label != "unknown" else ""}
                | Interactive controls: Zoom, Pan, Reset</em>
            </div>
        </div>
""")

        graphs_html_combined = '\n'.join(graphs_section)
        
        # Build complete HTML with separate Plotly.js CDN inclusion at top
        combined_html = f"""
<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <title>{title}</title>
    
    <!-- Include Plotly.js CDN once for all graphs -->
    <script src="https://cdn.plot.ly/plotly-2.26.0.min.js"></script>
    
    <style>
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            margin: 0;
            padding: 20px;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
        }}
        
        .container {{
            max-width: 1900px;
            margin: 0 auto;
            background-color: white;
            padding: 40px;
            border-radius: 16px;
            box-shadow: 0 10px 40px rgba(0,0,0,0.3);
        }}
        
        h1 {{
            color: #0071c5;
            text-align: center;
            font-size: 2.8em;
            margin-bottom: 40px;
            text-shadow: 2px 2px 4px rgba(0,0,0,0.1);
            border-bottom: 5px solid #0071c5;
            padding-bottom: 20px;
        }}
        
        .instrument-pane {{
            margin: 40px 0;
            padding: 25px;
            background: linear-gradient(135deg, #ffffff 0%, #f8f9fa 100%);
            border-radius: 12px;
            border: 2px solid #0071c5;
            box-shadow: 0 4px 6px rgba(0,0,0,0.1);
            transition: transform 0.2s, box-shadow 0.2s;
        }}
        
        .instrument-pane:hover {{
            transform: translateY(-5px);
            box-shadow: 0 8px 15px rgba(0,0,0,0.2);
        }}
        
        .pane-header {{
            display: flex;
            align-items: center;
            justify-content: space-between;
            margin-bottom: 20px;
            padding-bottom: 15px;
            border-bottom: 3px solid #0071c5;
        }}
        
        .plotly-graph-container {{
            background-color: white;
            padding: 20px;
            border-radius: 8px;
            border: 1px solid #e0e0e0;
            min-height: 600px;
        }}
        
        .pane-footer {{
            margin-top: 15px;
            padding-top: 10px;
            border-top: 1px solid #e0e0e0;
            text-align: right;
            color: #666;
            font-size: 0.85em;
        }}
        
        /* Scrollbar styling */
        ::-webkit-scrollbar {{
            width: 12px;
        }}
        
        ::-webkit-scrollbar-track {{
            background: #f1f1f1;
            border-radius: 10px;
        }}
        
        ::-webkit-scrollbar-thumb {{
            background: #0071c5;
            border-radius: 10px;
        }}
        
        ::-webkit-scrollbar-thumb:hover {{
            background: #005a9e;
        }}
    </style>
</head>
<body>
    <div class="container">
        <h1>🔋 {title}</h1>
        
        <!-- GENI Trend Analysis Section -->
        {geni_html if geni_html else ''}
        
        <!-- HSDES Sighting Query Section -->
        {hsdes_html if hsdes_html else ''}
        
        <!-- Instrument Graphs Section - Each in Separate Pane -->
        <div class="instruments-container">
            {graphs_html_combined}
        </div>
        
        <div style="margin-top: 60px; padding-top: 20px; border-top: 3px solid #0071c5; text-align: center; color: #666; font-size: 0.95em;">
            <p style="margin: 5px 0;"><strong>Generated by PowerKPI Dashboard Generator v3.4</strong></p>
            <p style="margin: 5px 0; font-size: 0.85em;">Multi-Instrument Analysis | Each graph is independently interactive</p>
            <p style="margin: 5px 0; font-size: 0.8em; color: #999;">© Intel Corporation</p>
        </div>
    </div>
</body>
</html>
"""
        
        # Write to file
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(combined_html)
        
        logger.info(f"Multi-instrument dashboard saved to: {output_path} with {len(graphs_html)} separate panes")
    
    # ------------------------------------------------------------------
    # Instrument-aware unit helpers
    # ------------------------------------------------------------------

    def _infer_unit_label(self, df: pd.DataFrame, instrument: str) -> str:
        """
        Infer a human-readable Y-axis unit label from the data.

        Because _split_by_unit is called before chart creation, df here is
        always unit-homogeneous.  Returns the single unit string or falls back
        to an instrument-level default.
        """
        if 'unit' in df.columns:
            units = [str(u).strip() for u in df['unit'].dropna().unique() if str(u).strip()]
            if len(units) == 1:
                return units[0]
            elif len(units) > 1:
                # Should not normally reach here after _split_by_unit, but keep safe
                return ' / '.join(sorted(set(units)))

        # Instrument-level fallback (only reached when 'unit' column absent)
        defaults = {
            'daq': 'W / A / V',
            'flexlogger': 'W / A / V',
            'socwatch': '% / MHz / ms',
            'perftracer': 'counts / ns',
            'powertrace': 'W / mW',
        }
        return defaults.get(instrument, 'Value')

    def _split_by_unit(self, df: pd.DataFrame) -> List[tuple]:
        """
        Split a DataFrame into groups where every row shares the same unit.

        Returns a list of (unit_label, sub_df) tuples, ordered so that the
        most-common unit appears first.  Rows with a blank/missing unit are
        grouped together under the label 'unknown'.

        This ensures we never mix e.g. '%' residency metrics with 'MHz'
        frequency metrics, or 'W' power with 'A' current, on the same chart.
        """
        if 'unit' not in df.columns or df.empty:
            return [('', df)]

        # Normalise: fill NaN / blank with 'unknown'
        df = df.copy()
        df['_unit_key'] = df['unit'].fillna('').astype(str).str.strip()
        df['_unit_key'] = df['_unit_key'].replace('', 'unknown')

        # Build groups ordered by frequency (most data points first)
        unit_counts = df['_unit_key'].value_counts()
        groups = []
        for unit_label in unit_counts.index:
            sub = df[df['_unit_key'] == unit_label].drop(columns=['_unit_key'])
            groups.append((unit_label, sub))

        return groups

    def _instrument_display_name(self, instrument: str) -> str:
        names = {
            'daq': 'DAQ — Power Rails',
            'flexlogger': 'DAQ — Power Rails',
            'socwatch': 'SocWatch — C-States & Frequencies',
            'perftracer': 'PerfTracer — Performance Metrics',
            'powertrace': 'PowerTrace — Power Trace',
        }
        return names.get(instrument, instrument.upper())

    # ------------------------------------------------------------------
    # Scatter plot — designed for DAQ / power rail data
    # Each point = one metric measurement.  X = metric name, Y = value.
    # One trace per workload so workloads can be compared side-by-side.
    # ------------------------------------------------------------------

    def create_scatter_plot(self, df: pd.DataFrame, workloads: List[str],
                            instrument: str = 'daq') -> go.Figure:
        """
        Scatter plot: X = power-rail / metric name, Y = measured value.

        Designed for DAQ/power data.  Each workload is a separate trace so
        rail values can be visually compared across runs.
        """
        fig = go.Figure()

        unit_label = self._infer_unit_label(df, instrument)
        inst_name = self._instrument_display_name(instrument)

        # Aggregate: one point per (workload, metric) — mean if repeated
        agg = df.groupby(['workload', 'metric'])['value_numeric'].mean().reset_index()

        # Collect all distinct metrics and sort them for a consistent X axis
        all_metrics = sorted(agg['metric'].unique())

        for workload in workloads:
            wl_agg = agg[agg['workload'] == workload]
            # Align to the full metric list so every workload shares the same X positions
            wl_dict = dict(zip(wl_agg['metric'], wl_agg['value_numeric']))
            y_vals = [wl_dict.get(m, None) for m in all_metrics]

            fig.add_trace(go.Scatter(
                name=workload,
                x=all_metrics,
                y=y_vals,
                mode='markers',
                marker=dict(size=10, opacity=0.8),
                hovertemplate=(
                    '<b>%{x}</b><br>'
                    f'Workload: {workload}<br>'
                    'Value: %{y:.4f} ' + unit_label +
                    '<extra></extra>'
                ),
            ))

        fig.update_layout(
            title=f'{inst_name} — Scatter Plot',
            xaxis_title='Power Rail / Metric',
            yaxis_title=f'Value ({unit_label})',
            xaxis=dict(tickangle=-45),
            template='plotly_white',
            height=700,
            hovermode='closest',
            legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='right', x=1),
        )

        return fig

    # ------------------------------------------------------------------
    # Bar chart — designed for SocWatch / PerfTracer / PowerTrace data
    # Each group on X = metric name; bars = workloads.
    # ------------------------------------------------------------------

    def create_bar_chart(self, df: pd.DataFrame, workloads: List[str],
                         instrument: str = 'socwatch') -> go.Figure:
        """
        Grouped bar chart: X = metric name, grouped bars per workload.

        Designed for SocWatch (C-states, residencies, frequencies) and
        similar categorical metrics.  Never mixed with power-rail data.
        """
        fig = go.Figure()

        unit_label = self._infer_unit_label(df, instrument)
        inst_name = self._instrument_display_name(instrument)

        agg = df.groupby(['workload', 'metric'])['value_numeric'].mean().reset_index()
        all_metrics = sorted(agg['metric'].unique())

        for workload in workloads:
            wl_agg = agg[agg['workload'] == workload]
            wl_dict = dict(zip(wl_agg['metric'], wl_agg['value_numeric']))
            y_vals = [wl_dict.get(m, None) for m in all_metrics]

            # Round for display; skip None
            text_vals = [f'{v:.3f}' if v is not None else '' for v in y_vals]

            fig.add_trace(go.Bar(
                name=workload,
                x=all_metrics,
                y=y_vals,
                text=text_vals,
                textposition='auto',
                hovertemplate=(
                    '<b>%{x}</b><br>'
                    f'Workload: {workload}<br>'
                    'Value: %{y:.4f} ' + unit_label +
                    '<extra></extra>'
                ),
            ))

        fig.update_layout(
            title=f'{inst_name} — Bar Chart',
            xaxis_title='Metric',
            yaxis_title=f'Value ({unit_label})',
            xaxis=dict(tickangle=-45),
            barmode='group',
            template='plotly_white',
            height=700,
            hovermode='closest',
            legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='right', x=1),
        )

        return fig

    def create_grouped_bar_chart(self, df: pd.DataFrame, workloads: List[str],
                                  instrument: str = 'socwatch') -> go.Figure:
        """Grouped bar chart — alias for create_bar_chart with explicit grouping note."""
        return self.create_bar_chart(df, workloads, instrument=instrument)

    # ------------------------------------------------------------------
    # Line graph — generic trend across metrics
    # ------------------------------------------------------------------

    def create_line_graph(self, df: pd.DataFrame, workloads: List[str],
                          instrument: str = 'daq') -> go.Figure:
        """Line graph: X = metric name (sorted), Y = value.  One line per workload."""
        fig = go.Figure()

        unit_label = self._infer_unit_label(df, instrument)
        inst_name = self._instrument_display_name(instrument)

        agg = df.groupby(['workload', 'metric'])['value_numeric'].mean().reset_index()
        all_metrics = sorted(agg['metric'].unique())

        for workload in workloads:
            wl_agg = agg[agg['workload'] == workload]
            wl_dict = dict(zip(wl_agg['metric'], wl_agg['value_numeric']))
            y_vals = [wl_dict.get(m, None) for m in all_metrics]

            fig.add_trace(go.Scatter(
                name=workload,
                x=all_metrics,
                y=y_vals,
                mode='lines+markers',
                marker=dict(size=8),
                line=dict(width=2),
                hovertemplate=(
                    '<b>%{x}</b><br>'
                    f'Workload: {workload}<br>'
                    'Value: %{y:.4f} ' + unit_label +
                    '<extra></extra>'
                ),
            ))

        fig.update_layout(
            title=f'{inst_name} — Line Graph',
            xaxis_title='Metric',
            yaxis_title=f'Value ({unit_label})',
            xaxis=dict(tickangle=-45),
            template='plotly_white',
            height=700,
            hovermode='x unified',
            legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='right', x=1),
        )

        return fig

    # ------------------------------------------------------------------
    # Heatmap
    # ------------------------------------------------------------------

    def create_heatmap(self, df: pd.DataFrame, workloads: List[str],
                       instrument: str = 'daq') -> go.Figure:
        """Heatmap: rows = metrics, columns = workloads, cells = mean value."""
        unit_label = self._infer_unit_label(df, instrument)
        inst_name = self._instrument_display_name(instrument)

        pivot_data = df.groupby(['metric', 'workload'])['value_numeric'].mean().reset_index()
        pivot_table = pivot_data.pivot(index='metric', columns='workload', values='value_numeric')

        fig = go.Figure(data=go.Heatmap(
            z=pivot_table.values,
            x=pivot_table.columns.tolist(),
            y=pivot_table.index.tolist(),
            colorscale='Viridis',
            hovertemplate=(
                'Metric: %{y}<br>Workload: %{x}<br>'
                'Value: %{z:.4f} ' + unit_label + '<extra></extra>'
            ),
        ))

        fig.update_layout(
            title=f'{inst_name} — Heatmap',
            xaxis_title='Workload',
            yaxis_title='Metric',
            template='plotly_white',
            height=max(700, len(pivot_table) * 22),
            width=max(900, len(workloads) * 160),
        )

        return fig

    # ------------------------------------------------------------------
    # Box plot
    # ------------------------------------------------------------------

    def create_box_plot(self, df: pd.DataFrame, workloads: List[str],
                        instrument: str = 'daq') -> go.Figure:
        """Box plot: distribution of values per workload."""
        fig = go.Figure()

        unit_label = self._infer_unit_label(df, instrument)
        inst_name = self._instrument_display_name(instrument)

        for workload in workloads:
            wl_data = df[df['workload'] == workload]

            fig.add_trace(go.Box(
                name=workload,
                y=wl_data['value_numeric'],
                boxmean='sd',
                hovertemplate='<b>%{text}</b><br>Value: %{y:.4f} ' + unit_label + '<extra></extra>',
                text=wl_data['metric'],
            ))

        fig.update_layout(
            title=f'{inst_name} — Box Plot',
            xaxis_title='Workload',
            yaxis_title=f'Value ({unit_label})',
            template='plotly_white',
            height=700,
            hovermode='x',
        )

        return fig


def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(description="Power KPI Dashboard Generator v2.0 (GUI)")
    parser.add_argument('--folder', type=Path, help='Initial folder to scan')
    
    args = parser.parse_args()
    
    root = tk.Tk()
    app = PowerKPIDashboardGUI(root, initial_folder=args.folder)
    root.mainloop()


if __name__ == "__main__":
    main()
