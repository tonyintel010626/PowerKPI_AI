import openpyxl
import statistics
import csv
import os

# ---- Load both workbooks ----
wb_sum = openpyxl.load_workbook("NVLP_WW15'26_Power_summary.xlsx", read_only=True, data_only=True)
wb_kpi = openpyxl.load_workbook("NVL P A0 PO BL KPIs WW15'26.xlsx", read_only=True, data_only=True)

# ---- Pre-silicon projections (Q1'26) from WW15_NVL P sheet ----
# Row 13 = header, rows 14-22 = workload KPI data
# Col C = Q1'26 Update (mW)
presil_kpi = {
    'IDON':     144,
    'CMS':       48,
    'Netflix':   728,
    'Youtube':  1223,
    'Catapult': 1120,
    'Busy Idle': 286,
}

# ---- Extract WW15DB Run2 (median) per workload from Power Summary ----
# WW15DB: cols F=RUN1, G=RUN2 (median), H=RUN3 (1-based: 6,7,8; 0-based: 5,6,7)
workload_sheets = ['IDON', 'CMS', 'Catapult', 'Youtube', 'Netflix']

def get_ww15_data(shname):
    ws = wb_sum[shname]
    data = {}
    for row in ws.iter_rows(min_row=11, max_row=24, min_col=1, max_col=8, values_only=True):
        if not row[0] or row[0] in ('Power Rails', 'Iteration'):
            continue
        rail = str(row[0]).strip()
        presil = row[1]
        ww15_r1 = row[5]
        ww15_r2 = row[6]  # RUN2 = median
        ww15_r3 = row[7]
        data[rail] = {
            'presil': presil,
            'ww15_run1': ww15_r1,
            'ww15_run2': ww15_r2,
            'ww15_run3': ww15_r3,
            'median': ww15_r2,
        }
    return data

results = {}
for sh in workload_sheets:
    results[sh] = get_ww15_data(sh)

# ---- Rail list in display order ----
rails_ordered = [
    'P_VCC_LP_ECORE', 'P_VCCPRIM_IO', 'P_VCCPRIM_VNNAON', 'P_VCCPRIM_1P8',
    'P_VCCRTC', 'P_VCCPRIM_1P8_FLTRA', 'P_VCCPRIM_1P8_FLTRB',
    'P_VDDQ_CPU', 'P_VDD2H_CPU', 'P_VCCCORE', 'P_VCCGT', 'P_VCCSA', 'P_MCP_TOTAL'
]

workloads_display = ['IDON', 'CMS', 'Catapult', 'Youtube', 'Netflix']
workload_labels = {
    'IDON': 'IDON (25C)',
    'CMS': 'CMS (25C)',
    'Catapult': 'Catapult @40C',
    'Youtube': 'Youtube @40C',
    'Netflix': 'Netflix @40C',
}

# ---- Compute ratios ----
def safe_ratio(post, pre):
    try:
        if pre and float(pre) != 0 and post is not None:
            return round(float(post) / float(pre), 3)
        return 'N/A'
    except:
        return 'N/A'

# ---- Build comparison rows ----
table_rows = []
for rail in rails_ordered:
    row = {'Power Rail': rail}
    for wl in workloads_display:
        rail_data = results.get(wl, {}).get(rail, {})
        presil = rail_data.get('presil')
        post = rail_data.get('median')
        ratio = safe_ratio(post, presil)
        lbl = workload_labels[wl]
        row[f'{lbl} Pre (mW)'] = round(float(presil), 3) if presil is not None else 'N/A'
        row[f'{lbl} Post WW15 (mW)'] = round(float(post), 3) if post is not None else 'N/A'
        row[f'{lbl} Ratio'] = ratio
    table_rows.append(row)

# ---- System-level KPI comparison table ----
# From WW15_NVL P sheet: measured WW10 is the latest run (col E)
# For WW15 measurement we use P_MCP_TOTAL from power summary
kpi_table = []
kpi_workloads = [
    ('IDON',     'OS Idle (25C)',          144),
    ('CMS',      'MCS (25C)',               48),
    ('Netflix',  'Netflix 1080p AVC (40C)', 728),
    ('Youtube',  'AVI 4K YouTube (40C)',   1223),
    ('Catapult', '4-tab Catapult (40C)',   1120),
]
for wl, label, presil_mw in kpi_workloads:
    mcp = results.get(wl, {}).get('P_MCP_TOTAL', {})
    post_median = mcp.get('median')
    ratio = safe_ratio(post_median, presil_mw)
    kpi_table.append({
        'Workload': label,
        'Pre-Silicon Q1\'26 (mW)': presil_mw,
        'Post-Silicon WW15 Median (mW)': round(float(post_median), 2) if post_median is not None else 'N/A',
        'Ratio (Post/Pre)': ratio,
        'Status': 'OVER' if isinstance(ratio, float) and ratio > 1.0 else ('OK' if isinstance(ratio, float) else 'N/A'),
    })

# ---- Build CSV ----
csv_path = 'NVL_P_WW15_PrePost_Report.csv'

# Section 1: KPI summary
with open(csv_path, 'w', newline='') as f:
    writer = csv.writer(f)
    writer.writerow(['=== NVL-P WW15 Power KPI: Pre-Silicon (Q1\'26) vs Post-Silicon (WW15 Median) ==='])
    writer.writerow([])
    writer.writerow(['SYSTEM-LEVEL KPI COMPARISON (P_MCP_TOTAL)'])
    kpi_cols = list(kpi_table[0].keys())
    writer.writerow(kpi_cols)
    for r in kpi_table:
        writer.writerow([r[c] for c in kpi_cols])
    writer.writerow([])
    writer.writerow(['POWER RAIL BREAKDOWN (WW15 Run2 = Median)'])
    cols = list(table_rows[0].keys())
    writer.writerow(cols)
    for r in table_rows:
        writer.writerow([r.get(c, '') for c in cols])

print(f"CSV saved: {csv_path}")

# ---- Build HTML ----
def color_ratio(ratio):
    if not isinstance(ratio, (int, float)):
        return '#aaaaaa'
    if ratio <= 1.0:
        return '#27ae60'
    elif ratio <= 1.3:
        return '#f39c12'
    else:
        return '#e74c3c'

def ratio_badge(ratio):
    c = color_ratio(ratio)
    val = f'{ratio:.3f}' if isinstance(ratio, float) else str(ratio)
    return f'<span style="background:{c};color:white;padding:2px 6px;border-radius:4px;font-weight:bold">{val}</span>'

# Build header for rail table
def build_rail_table_html():
    wl_col_groups = []
    for wl in workloads_display:
        lbl = workload_labels[wl]
        wl_col_groups.append(lbl)

    html = '<table border="1" cellspacing="0" cellpadding="6" style="border-collapse:collapse;font-size:12px;width:100%">'
    # Multi-level header
    html += '<thead>'
    html += '<tr style="background:#2c3e50;color:white">'
    html += '<th rowspan="2" style="min-width:140px">Power Rail</th>'
    for lbl in wl_col_groups:
        html += f'<th colspan="3" style="text-align:center">{lbl}</th>'
    html += '</tr>'
    html += '<tr style="background:#34495e;color:white">'
    for lbl in wl_col_groups:
        html += '<th>Pre (mW)</th><th>Post WW15 (mW)</th><th>Ratio</th>'
    html += '</tr>'
    html += '</thead><tbody>'

    for i, row in enumerate(table_rows):
        bg = '#f9f9f9' if i % 2 == 0 else 'white'
        is_total = row['Power Rail'] == 'P_MCP_TOTAL'
        row_style = f'background:{bg};font-weight:{"bold" if is_total else "normal"}'
        html += f'<tr style="{row_style}">'
        html += f'<td style="font-family:monospace">{row["Power Rail"]}</td>'
        for wl in workloads_display:
            lbl = workload_labels[wl]
            pre = row.get(f'{lbl} Pre (mW)', 'N/A')
            post = row.get(f'{lbl} Post WW15 (mW)', 'N/A')
            ratio = row.get(f'{lbl} Ratio', 'N/A')
            html += f'<td style="text-align:right">{pre}</td>'
            html += f'<td style="text-align:right">{post}</td>'
            html += f'<td style="text-align:center">{ratio_badge(ratio)}</td>'
        html += '</tr>'

    html += '</tbody></table>'
    return html

def build_kpi_table_html():
    html = '<table border="1" cellspacing="0" cellpadding="8" style="border-collapse:collapse;font-size:13px;width:700px">'
    html += '<thead><tr style="background:#1a252f;color:white">'
    for col in ['Workload', "Pre-Silicon Q1'26 (mW)", 'Post-Silicon WW15 Median (mW)', 'Ratio (Post/Pre)', 'Status']:
        html += f'<th>{col}</th>'
    html += '</tr></thead><tbody>'
    for i, row in enumerate(kpi_table):
        bg = '#f2f8ff' if i % 2 == 0 else 'white'
        ratio = row['Ratio (Post/Pre)']
        status = row['Status']
        status_color = '#27ae60' if status == 'OK' else '#e74c3c'
        html += f'<tr style="background:{bg}">'
        html += f'<td><b>{row["Workload"]}</b></td>'
        html += f'<td style="text-align:right">{row["Pre-Silicon Q1\'26 (mW)"]}</td>'
        html += f'<td style="text-align:right;font-weight:bold">{row["Post-Silicon WW15 Median (mW)"]}</td>'
        html += f'<td style="text-align:center">{ratio_badge(ratio)}</td>'
        html += f'<td style="text-align:center;color:{status_color};font-weight:bold">{status}</td>'
        html += '</tr>'
    html += '</tbody></table>'
    return html

# Build summary narrative
def build_summary():
    lines = []
    lines.append('<h3>Executive Summary</h3>')
    lines.append('<ul>')
    for row in kpi_table:
        wl = row['Workload']
        post = row['Post-Silicon WW15 Median (mW)']
        pre = row["Pre-Silicon Q1'26 (mW)"]
        ratio = row['Ratio (Post/Pre)']
        if isinstance(ratio, float):
            pct = round((ratio - 1.0) * 100, 1)
            direction = f'+{pct}% above' if pct > 0 else f'{abs(pct)}% below'
            color = '#e74c3c' if pct > 0 else '#27ae60'
            lines.append(f'<li><b>{wl}</b>: Post={post} mW vs Pre={pre} mW &mdash; '
                         f'<span style="color:{color}">{direction} projection</span> (ratio={ratio})</li>')
        else:
            lines.append(f'<li><b>{wl}</b>: No data available</li>')
    lines.append('</ul>')
    return '\n'.join(lines)

html_out = f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<title>NVL-P Power KPI Report: Pre vs Post Silicon (WW15)</title>
<style>
  body {{ font-family: Arial, sans-serif; margin: 30px; color: #222; }}
  h1 {{ color: #1a252f; border-bottom: 3px solid #2980b9; padding-bottom: 8px; }}
  h2 {{ color: #2c3e50; margin-top: 30px; }}
  h3 {{ color: #34495e; }}
  .metadata {{ background:#ecf0f1; padding:12px; border-radius:6px; margin-bottom:20px; font-size:13px; }}
  .section {{ margin-top: 30px; }}
  table {{ margin-bottom: 20px; }}
  .legend {{ margin: 10px 0; font-size: 12px; }}
  .legend span {{ display:inline-block; width:18px; height:14px; margin-right:5px; vertical-align:middle; border-radius:3px; }}
</style>
</head>
<body>
<h1>NVL-P A0 Power KPI Report &mdash; Pre-Silicon vs Post-Silicon</h1>
<div class="metadata">
  <b>Pre-Silicon Reference:</b> Q1&apos;26 Projection (from NVL P A0 PO BL KPIs WW15&apos;26.xlsx)<br>
  <b>Post-Silicon Data:</b> WW15 Run2 (Median) from NVLP_WW15&apos;26_Power_summary.xlsx<br>
  <b>Unit:</b> milliwatts (mW) &mdash; Summary covers G12:G24 (P_MCP_TOTAL row is system-level total)<br>
  <b>Note:</b> Netflix WW15 data = 0 (no measurement available yet)
</div>

<div class="legend">
  <b>Ratio color code:</b>
  <span style="background:#27ae60"></span> &le;1.0 (meets/beats projection)
  &nbsp;
  <span style="background:#f39c12"></span> 1.0&ndash;1.3 (moderately over)
  &nbsp;
  <span style="background:#e74c3c"></span> &gt;1.3 (significantly over projection)
</div>

<div class="section">
<h2>1. System-Level KPI Summary (P_MCP_TOTAL)</h2>
{build_kpi_table_html()}
</div>

<div class="section">
{build_summary()}
</div>

<div class="section">
<h2>2. Power Rail Breakdown (WW15 Run2 = Median)</h2>
<p style="font-size:12px;color:#666">Pre-silicon projections from NVL_P_Q1&apos;26 tab. Post-silicon = WW15 Run2 median value per workload.</p>
{build_rail_table_html()}
</div>

<div class="section">
<h2>3. Brief Summary &mdash; Power Ratio Analysis</h2>
<table border="1" cellpadding="8" cellspacing="0" style="border-collapse:collapse;font-size:13px">
<thead><tr style="background:#2c3e50;color:white">
  <th>Workload</th><th>Post/Pre Ratio</th><th>Interpretation</th>
</tr></thead>
<tbody>
"""
for row in kpi_table:
    ratio = row['Ratio (Post/Pre)']
    wl = row['Workload']
    if isinstance(ratio, float):
        if ratio <= 1.0:
            interp = 'Meets or beats pre-silicon projection'
            bg = '#eafaf1'
        elif ratio <= 1.3:
            interp = 'Moderately above projection — monitor'
            bg = '#fef9e7'
        elif ratio <= 2.0:
            interp = 'Significantly above projection — investigate'
            bg = '#fdf2f8'
        else:
            interp = 'Critically above projection — action required'
            bg = '#fdedec'
    else:
        interp = 'No measurement data'
        bg = '#f4f4f4'
    html_out += f'<tr style="background:{bg}"><td><b>{wl}</b></td><td style="text-align:center">{ratio_badge(ratio)}</td><td>{interp}</td></tr>\n'

html_out += """</tbody></table>
</div>

<div class="section">
<h2>4. Notes</h2>
<ul>
  <li>Netflix WW15 data shows all zeros — measurement not yet captured for WW15 (only WW15 header row present, runs TBD).</li>
  <li>IDON, CMS, Catapult, Youtube show WW15 Run2 as the representative median run per test design.</li>
  <li>P_VCCPRIM_VNNAON consistently above pre-silicon projection across workloads — potential VNNAON leakage issue.</li>
  <li>P_VCCSA is the dominant contributor to excess power in Youtube and Catapult workloads.</li>
</ul>
</div>

</body>
</html>
"""

html_path = 'NVL_P_WW15_PrePost_Report.html'
with open(html_path, 'w', encoding='utf-8') as f:
    f.write(html_out)

print(f"HTML report saved: {html_path}")
print(f"CSV report saved: {csv_path}")
print("\nKPI Summary:")
for row in kpi_table:
    pre_key = "Pre-Silicon Q1'26 (mW)"
    print(f"  {row['Workload']}: pre={row[pre_key]}, post={row['Post-Silicon WW15 Median (mW)']}, ratio={row['Ratio (Post/Pre)']}")
